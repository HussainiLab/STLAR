#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Batch processing script for Spatial Spectral Mapper (SSM)
Allows command-line analysis of LFP data without GUI
"""

import os
import sys
import argparse
import csv
import numpy as np
from pathlib import Path
import gc  # Garbage collection for memory management
import psutil  # For memory monitoring
import subprocess

# Set matplotlib to non-interactive backend BEFORE importing anything else
import matplotlib
from core.processors.spectral_functions import (
    export_binned_analysis_to_csv,
    visualize_binned_analysis,
    export_binned_analysis_jpgs,
    BATCH_CMAP_POWER,
    BATCH_CMAP_PERCENT,
    BATCH_CMAP_OCCUPANCY,
    BATCH_CMAP_DOMINANT,
)
matplotlib.use('Agg')  # Use Agg backend - no display needed
from matplotlib import pyplot as plt
from matplotlib.patches import Ellipse
import os
os.environ['QT_QPA_PLATFORM'] = 'offscreen'  # Set Qt to offscreen mode

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

# Now import Qt with QGuiApplication (offscreen mode)
from PyQt5.QtGui import QGuiApplication
app = QGuiApplication(sys.argv)  # Create app in offscreen mode

from src.initialize_fMap import initialize_fMap
from src.core.data_loaders import grab_position_data
from src.core.processors.Tint_Matlab import speed2D
from src.core.processors.spectral_functions import speed_bins


class BatchWorker:
    """Worker class for batch processing without GUI"""
    
    class MockSignals:
        """Mock signal object for batch processing (replaces Qt signals)"""
        def __init__(self):
            self.progress_value = 0
            self.progress_text = ""
        
        def emit(self, value=None):
            """Mock signal emit - just log the value"""
            if isinstance(value, str):
                self.progress_text = value
                print(f"  → {value}")
            else:
                self.progress_value = value
    
    def __init__(self, output_dir=None):
        self.output_dir = output_dir
        self.progress_messages = []
        self.signals = self.MockSignals()
        
        # Create nested structure for mock signals
        class ProgressSignals:
            def __init__(self, worker_signals):
                self.worker_signals = worker_signals
            
            def emit(self, value):
                if isinstance(value, str):
                    self.worker_signals.progress_text = value
                    print(f"  → {value}")
                else:
                    self.worker_signals.progress_value = value
        
        class TextProgressSignals:
            def __init__(self, worker_signals):
                self.worker_signals = worker_signals
            
            def emit(self, value):
                self.worker_signals.progress_text = value
                print(f"  → {value}")
        
        self.signals.progress = ProgressSignals(self.signals)
        self.signals.text_progress = TextProgressSignals(self.signals)
    
    def log(self, message):
        """Log progress messages"""
        print(f"[SSM] {message}")
        self.progress_messages.append(message)


def find_pos_file(eeg_file):
    """Auto-detect .pos file based on .eeg/.egf filename"""
    base_name = os.path.splitext(eeg_file)[0]
    pos_file = base_name + '.pos'
    
    if os.path.exists(pos_file):
        return pos_file
    else:
        raise FileNotFoundError(f"Position file not found: {pos_file}")


def find_recording_files(directory):
    """
    Find all EEG/EGF recording files in a directory.
    Returns list of (egf/eeg file, pos file) tuples.
    
    Priority: Use .egf if exists, otherwise .eeg
    Skip numbered variants (egf2-4, eeg2-4) if base file exists
    """
    import glob
    from pathlib import Path
    
    recordings = {}  # basename -> (electrophys_file, pos_file)
    
    # Find all .egf and .eeg files
    egf_files = glob.glob(os.path.join(directory, "*.egf"))
    eeg_files = glob.glob(os.path.join(directory, "*.eeg"))
    
    # Process EGF files first (higher priority)
    for egf_file in egf_files:
        basename = Path(egf_file).stem
        # Skip numbered variants (egf2, egf3, egf4)
        if basename.endswith(('2', '3', '4')) and basename[:-1] in [Path(f).stem for f in egf_files]:
            continue
        
        try:
            pos_file = find_pos_file(egf_file)
            recordings[basename] = (egf_file, pos_file)
        except FileNotFoundError:
            print(f"⚠ Skipping {os.path.basename(egf_file)} - no .pos file found")
    
    # Process EEG files (only if no EGF exists for same basename)
    for eeg_file in eeg_files:
        basename = Path(eeg_file).stem
        
        # Skip if already have EGF for this basename
        if basename in recordings:
            continue
        
        # Skip numbered variants (eeg2, eeg3, eeg4)
        if basename.endswith(('2', '3', '4')) and basename[:-1] in [Path(f).stem for f in eeg_files]:
            continue
        
        try:
            pos_file = find_pos_file(eeg_file)
            recordings[basename] = (eeg_file, pos_file)
        except FileNotFoundError:
            print(f"⚠ Skipping {os.path.basename(eeg_file)} - no .pos file found")
    
    return list(recordings.values())


def find_eoi_file(electrophys_file):
    """Find associated EOI file (CSV or HFOScores)"""
    base_path = os.path.splitext(electrophys_file)[0]
    base_name = os.path.basename(base_path)
    parent_dir = os.path.dirname(electrophys_file)
    
    # 1. Check for {basename}_EOI.csv
    csv_candidate = f"{base_path}_EOI.csv"
    if os.path.exists(csv_candidate):
        return csv_candidate
        
    # 2. Check for HFOScores/{basename}/{basename}_HIL.txt (or similar)
    # Try standard HFO detection outputs
    for tag in ['HIL', 'STE', 'MNI', 'DL', 'CON']:
        score_path = os.path.join(parent_dir, 'HFOScores', base_name, f"{base_name}_{tag}.txt")
        if os.path.exists(score_path):
            return score_path
            
    return None


def resolve_eoi_file(electrophys_file, eoi_file_override=None):
    """Resolve EOI file from explicit path or auto-detection."""
    if not eoi_file_override:
        return find_eoi_file(electrophys_file)

    override = os.path.expanduser(eoi_file_override)
    if os.path.isdir(override):
        base_name = os.path.splitext(os.path.basename(electrophys_file))[0]
        candidates = [
            os.path.join(override, f"{base_name}_EOI.csv"),
            os.path.join(override, f"{base_name}.csv"),
            os.path.join(override, base_name, f"{base_name}_EOI.csv"),
            os.path.join(override, base_name, f"{base_name}.csv"),
            os.path.join(override, "HFOScores", base_name, f"{base_name}_EOI.csv"),
            os.path.join(override, "HFOScores", base_name, f"{base_name}.csv"),
        ]
        for tag in ['HIL', 'STE', 'MNI', 'DL', 'CON']:
            candidates.extend([
                os.path.join(override, f"{base_name}_{tag}.txt"),
                os.path.join(override, base_name, f"{base_name}_{tag}.txt"),
                os.path.join(override, "HFOScores", base_name, f"{base_name}_{tag}.txt"),
            ])
        for candidate in candidates:
            if os.path.exists(candidate):
                return candidate
        return None

    return override if os.path.exists(override) else None

def load_eois(eoi_file):
    """Load EOIs from CSV or score file. Returns list of dicts with start/stop (s) and label."""
    eois = []
    header_map = None
    try:
        with open(eoi_file, 'r') as f:
            reader = csv.reader(f, delimiter='\t' if eoi_file.endswith('.txt') else ',')
            for row in reader:
                if not row:
                    continue

                if header_map is None:
                    lowered = [cell.strip().lower() for cell in row]
                    if any("start" in cell for cell in lowered) and any("stop" in cell for cell in lowered):
                        header_map = {
                            "start": next(i for i, cell in enumerate(lowered) if "start" in cell),
                            "stop": next(i for i, cell in enumerate(lowered) if "stop" in cell),
                            "label": next((i for i, cell in enumerate(lowered) if "label" in cell or "band" in cell), None),
                        }
                        continue

                try:
                    if header_map is not None:
                        start = float(row[header_map["start"]])
                        stop = float(row[header_map["stop"]])
                        label = None
                        if header_map["label"] is not None and header_map["label"] < len(row):
                            label = row[header_map["label"]].strip() or None
                        if eoi_file.endswith('.txt'):
                            start /= 1000.0
                            stop /= 1000.0
                        eois.append({"start": start, "stop": stop, "label": label})
                        continue

                    # HFO score format: ID, Start(ms), Stop(ms), Settings, Label, ...
                    if eoi_file.endswith('.txt') and len(row) >= 3:
                        start = float(row[1]) / 1000.0
                        stop = float(row[2]) / 1000.0
                        label = row[4].strip() if len(row) >= 5 and row[4].strip() else None
                        eois.append({"start": start, "stop": stop, "label": label})
                    elif len(row) >= 2:
                        # Simple CSV: Start(s), Stop(s), optional Label
                        start = float(row[0])
                        stop = float(row[1])
                        label = row[2].strip() if len(row) >= 3 and row[2].strip() else None
                        eois.append({"start": start, "stop": stop, "label": label})
                except (ValueError, IndexError, StopIteration):
                    continue
    except Exception as e:
        print(f"  ⚠ Error loading EOIs from {eoi_file}: {e}")
    return eois


def _normalize_label(label):
    if not label:
        return "eoi"
    return " ".join(label.strip().lower().split())


def _label_color(label, fallback_index):
    palette = {
        "ripple": "#1f77b4",
        "fast ripple": "#ff7f0e",
        "fast_ripple": "#ff7f0e",
        "fr": "#ff7f0e",
        "artifact": "#d62728",
        "noise": "#d62728",
        "eoi": "#2ca02c",
        "unknown": "#7f7f7f",
    }
    normalized = _normalize_label(label)
    if normalized in palette:
        return palette[normalized]
    cycle = plt.get_cmap("tab10").colors
    return cycle[fallback_index % len(cycle)]


def plot_trajectory_with_eois(output_path, pos_x_chunks, pos_y_chunks, eoi_segments, ppm=None, arena_shape=None, binned_data=None):
    if pos_x_chunks is None or pos_y_chunks is None:
        print("  ⚠ No tracking data available for trajectory plot.")
        return

    def _flatten(chunks):
        arrays = [np.asarray(chunk) for chunk in chunks if len(chunk) > 0]
        return np.concatenate(arrays) if arrays else np.array([])

    def _to_cm(values):
        if ppm is not None and ppm > 0:
            return (values / ppm) * 100.0
        return values

    def _draw_bins(ax, min_x, max_x, min_y, max_y):
        is_polar = False
        if arena_shape:
            is_polar = ("Circle" in arena_shape or "Ellipse" in arena_shape)

        if is_polar:
            width = max_x - min_x
            height = max_y - min_y
            center_x = min_x + width / 2.0
            center_y = min_y + height / 2.0

            if width <= 0 or height <= 0:
                return

            e1 = Ellipse((center_x, center_y), width, height, fill=False, edgecolor='gray', linestyle='--', alpha=0.5)
            ax.add_patch(e1)
            scale = 1.0 / np.sqrt(2)
            e2 = Ellipse((center_x, center_y), width * scale, height * scale, fill=False, edgecolor='gray', linestyle='--', alpha=0.5)
            ax.add_patch(e2)

            angles = np.linspace(-np.pi, np.pi, 9)
            for theta in angles:
                x_edge = center_x + (width / 2.0) * np.cos(theta)
                y_edge = center_y + (height / 2.0) * np.sin(theta)
                ax.plot([center_x, x_edge], [center_y, y_edge], color='gray', linestyle='--', alpha=0.5, linewidth=0.8)
        else:
            if binned_data and 'x_bin_edges' in binned_data and 'y_bin_edges' in binned_data:
                x_edges = _to_cm(np.asarray(binned_data['x_bin_edges']))
                y_edges = _to_cm(np.asarray(binned_data['y_bin_edges']))
            else:
                x_edges = np.linspace(min_x, max_x, 5)
                y_edges = np.linspace(min_y, max_y, 5)
            for x in x_edges:
                ax.axvline(x, color='gray', linestyle='--', alpha=0.5, linewidth=0.8)
            for y in y_edges:
                ax.axhline(y, color='gray', linestyle='--', alpha=0.5, linewidth=0.8)

    def _compute_occupancy(traj_x, traj_y, min_x, max_x, min_y, max_y):
        is_polar = False
        if arena_shape:
            is_polar = ("Circle" in arena_shape or "Ellipse" in arena_shape)

        if traj_x.size == 0 or traj_y.size == 0:
            return None

        if is_polar:
            width = max_x - min_x
            height = max_y - min_y
            if width <= 0 or height <= 0:
                return None
            center_x = min_x + width / 2.0
            center_y = min_y + height / 2.0
            dx = traj_x - center_x
            dy = traj_y - center_y
            rx = width / 2.0 if width > 0 else 1.0
            ry = height / 2.0 if height > 0 else 1.0
            r_norm = np.sqrt((dx / rx) ** 2 + (dy / ry) ** 2)
            theta = np.arctan2(dy, dx)
            r_edges = np.array([0.0, 1.0 / np.sqrt(2), 1e9])
            theta_edges = np.linspace(-np.pi, np.pi, 9)
            occ_counts, _, _ = np.histogram2d(r_norm, theta, bins=[r_edges, theta_edges])
            total = np.sum(occ_counts)
            if total <= 0:
                return None
            occ_pct = (occ_counts / total) * 100.0
            return {
                "is_polar": True,
                "values": occ_pct,
                "radius_edges": np.array([0.0, 1.0 / np.sqrt(2), 1.0]),
                "angle_edges": theta_edges,
                "center": (center_x, center_y),
                "width": width,
                "height": height,
            }

        if binned_data and 'x_bin_edges' in binned_data and 'y_bin_edges' in binned_data:
            x_edges = _to_cm(np.asarray(binned_data['x_bin_edges']))
            y_edges = _to_cm(np.asarray(binned_data['y_bin_edges']))
        else:
            x_edges = np.linspace(min_x, max_x, 5)
            y_edges = np.linspace(min_y, max_y, 5)

        occ_counts, _, _ = np.histogram2d(traj_x, traj_y, bins=[x_edges, y_edges])
        total = np.sum(occ_counts)
        if total <= 0:
            return None
        occ_pct = (occ_counts / total) * 100.0
        x_centers = 0.5 * (x_edges[:-1] + x_edges[1:])
        y_centers = 0.5 * (y_edges[:-1] + y_edges[1:])
        return {
            "is_polar": False,
            "values": occ_pct,
            "x_centers": x_centers,
            "y_centers": y_centers,
        }

    def _add_occupancy_labels(ax, occ_info):
        if not occ_info or "values" not in occ_info:
            return
        values = occ_info["values"]
        if np.sum(values) <= 0:
            return
        if occ_info.get("is_polar"):
            center_x, center_y = occ_info["center"]
            width = occ_info["width"]
            height = occ_info["height"]
            r_edges = occ_info["radius_edges"]
            theta_edges = occ_info["angle_edges"]
            r_centers = 0.5 * (r_edges[:-1] + r_edges[1:])
            theta_centers = 0.5 * (theta_edges[:-1] + theta_edges[1:])
            for i, r in enumerate(r_centers):
                for j, theta in enumerate(theta_centers):
                    pct = values[i, j]
                    if pct <= 0:
                        continue
                    r_clamped = min(r, 1.0)
                    x = center_x + (width / 2.0) * r_clamped * np.cos(theta)
                    y = center_y + (height / 2.0) * r_clamped * np.sin(theta)
                    ax.text(x, y, f"{pct:.0f}%", fontsize=7, ha='center', va='center', color='#222',
                            bbox=dict(boxstyle='round,pad=0.15', facecolor='white', alpha=0.65, linewidth=0))
        else:
            x_centers = occ_info["x_centers"]
            y_centers = occ_info["y_centers"]
            for i, xc in enumerate(x_centers):
                for j, yc in enumerate(y_centers):
                    pct = values[i, j]
                    if pct <= 0:
                        continue
                    ax.text(xc, yc, f"{pct:.0f}%", fontsize=7, ha='center', va='center', color='#222',
                            bbox=dict(boxstyle='round,pad=0.15', facecolor='white', alpha=0.65, linewidth=0))

    traj_x = _flatten(pos_x_chunks)
    traj_y = _flatten(pos_y_chunks)

    traj_x = _to_cm(traj_x)
    traj_y = _to_cm(traj_y)
    unit = "cm" if ppm is not None and ppm > 0 else "px"

    if traj_x.size and traj_y.size:
        min_x, max_x = float(np.min(traj_x)), float(np.max(traj_x))
        min_y, max_y = float(np.min(traj_y)), float(np.max(traj_y))
    else:
        min_x, max_x, min_y, max_y = 0.0, 1.0, 0.0, 1.0

    fig, ax = plt.subplots(figsize=(8, 8))
    if traj_x.size and traj_y.size:
        ax.plot(traj_x, traj_y, color="#b0b0b0", linewidth=0.7, alpha=0.7, label="Trajectory")

    _draw_bins(ax, min_x, max_x, min_y, max_y)
    occ_info = _compute_occupancy(traj_x, traj_y, min_x, max_x, min_y, max_y)
    _add_occupancy_labels(ax, occ_info)

    used_labels = set()
    label_index = 0
    for chunk_idx, segments in eoi_segments.items():
        for seg_x, seg_y, label in segments:
            seg_x = _to_cm(np.asarray(seg_x))
            seg_y = _to_cm(np.asarray(seg_y))
            color = _label_color(label, label_index)
            display_label = label or "EOI"
            if display_label not in used_labels:
                ax.scatter(seg_x, seg_y, s=6, color=color, alpha=0.9, label=display_label)
                used_labels.add(display_label)
                label_index += 1
            else:
                ax.scatter(seg_x, seg_y, s=6, color=color, alpha=0.9)

    ax.set_title("Trajectory with EOIs")
    ax.set_xlabel(f"X ({unit})")
    ax.set_ylabel(f"Y ({unit})")
    ax.set_aspect("equal", adjustable="box")
    ax.legend(loc="best", fontsize=8)
    ax.grid(True, alpha=0.2)
    fig.tight_layout()
    fig.savefig(output_path, dpi=200, bbox_inches="tight")
    plt.close(fig)

def export_to_csv(output_path, pos_t, chunk_powers_data, chunk_size, ppm, 
                  pos_x_chunks, pos_y_chunks):
    """Export analysis results to Excel (or CSV fallback)"""
    
    # Try to use Excel; fallback to CSV
    try:
        import openpyxl
        use_excel = True
    except ImportError:
        use_excel = False
    
    bands = [
        ("Delta", "Avg Delta Power"),
        ("Theta", "Avg Theta Power"),
        ("Beta", "Avg Beta Power"),
        ("Low Gamma", "Avg Low Gamma Power"),
        ("High Gamma", "Avg High Gamma Power"),
        ("Ripple", "Avg Ripple Power"),
        ("Fast Ripple", "Avg Fast Ripple Power"),
    ]
    
    band_labels = [label for _, label in bands]
    percent_labels = [f"Percent {name}" for name, _ in bands]
    
    # Calculate distances
    distances_per_bin = []
    cumulative_distances = []
    cumulative_sum = 0.0
    
    if pos_x_chunks is not None and pos_y_chunks is not None:
        for i in range(len(pos_x_chunks)):
            distance_cm_in_bin = 0.0
            
            x_bin = pos_x_chunks[i]
            y_bin = pos_y_chunks[i]
            
            if len(x_bin) > 1:
                dx = np.diff(np.array(x_bin))
                dy = np.diff(np.array(y_bin))
                distances_in_bin_pixels = np.sqrt(dx**2 + dy**2)
                total_distance_pixels_in_bin = np.sum(distances_in_bin_pixels)
                
                if ppm is not None and ppm > 0:
                    distance_cm_in_bin = (total_distance_pixels_in_bin / ppm) * 100
                else:
                    distance_cm_in_bin = total_distance_pixels_in_bin
            
            distances_per_bin.append(distance_cm_in_bin)
            cumulative_sum += distance_cm_in_bin
            cumulative_distances.append(cumulative_sum)
    
    # Determine max full chunks
    actual_duration = float(pos_t[-1])
    max_full_chunks = int(actual_duration / chunk_size)
    num_rows = min(len(pos_t), max_full_chunks)
    
    # Gather per-band arrays
    band_arrays = {}
    for key, label in bands:
        arr = np.array(chunk_powers_data.get(key, [])).reshape(-1)
        band_arrays[label] = arr
    
    # Prepare header and data rows
    header = ["Time Bin (s)", "Distance Per Bin (cm)", "Cumulative Distance (cm)"] + band_labels + percent_labels
    data_rows = []
    
    for i in range(num_rows):
        time_bin_start = i * chunk_size
        time_bin_end = (i + 1) * chunk_size
        time_bin_str = f"{time_bin_start}-{time_bin_end}"
        
        row = [time_bin_str]
        
        # Add distance per bin
        if distances_per_bin and i < len(distances_per_bin):
            row.append(round(distances_per_bin[i], 3))
        else:
            row.append("")
        
        # Add cumulative distance
        if cumulative_distances and i < len(cumulative_distances):
            row.append(round(cumulative_distances[i], 3))
        else:
            row.append("")
        
        band_values = []
        for _, label in bands:
            val = band_arrays[label][i] if i < len(band_arrays[label]) else ""
            band_values.append(val)
        
        row.extend(band_values)
        
        numeric_vals = [float(v) for v in band_values if v != ""]
        total_power = sum(numeric_vals)
        
        for v in band_values:
            if v == "" or total_power == 0:
                row.append("")
            else:
                row.append(round((float(v) / total_power) * 100.0, 3))
        
        data_rows.append(row)
    
    # Write Excel or CSV
    if use_excel:
        output_file = output_path.replace('.csv', '.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'SSM Data'
        ws.append(header)
        for row in data_rows:
            ws.append(row)
        wb.save(output_file)
        print(f"✓ Excel exported to: {output_file}")
    else:
        with open(output_path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(header)
            for row in data_rows:
                writer.writerow(row)
        print(f"✓ CSV exported to: {output_path} (openpyxl not installed)")


def main():
    parser = argparse.ArgumentParser(
        description="Batch process LFP data with Spatial Spectral Mapper (SSM)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Process a single file
  python batch_ssm.py data/recording.eeg
  
  # Process all files in a directory
  python batch_ssm.py data/recordings/
  
  # With custom parameters
  python batch_ssm.py data/recording.eeg --ppm 500 --chunk-size 5 --speed-min 5 --speed-max 20
  
  # Batch directory with custom output
  python batch_ssm.py data/recordings/ --ppm 600 --speed-min 0 --speed-max 100 -o results/
        """
    )
    
    parser.add_argument(
        "input_path",
        help="Path to .eeg/.egf file OR directory containing multiple recordings"
    )
    
    parser.add_argument(
        "-o", "--output",
        default=None,
        help="Output directory for CSV file(s) (default: same directory as input file(s))"
    )
    
    parser.add_argument(
        "--ppm",
        type=int,
        default=600,
        help="Pixels per meter for position data (default: 600)"
    )
    
    parser.add_argument(
        "--chunk-size",
        type=int,
        default=60,
        help="Chunk size in seconds for analysis (default: 60)"
    )
    
    parser.add_argument(
        "--speed-min",
        type=float,
        default=0.0,
        help="Minimum speed threshold in cm/s (default: 0.0)"
    )
    
    parser.add_argument(
        "--speed-max",
        type=float,
        default=100.0,
        help="Maximum speed threshold in cm/s (default: 100.0)"
    )
    
    parser.add_argument(
        "--window",
        type=str,
        default="hann",
        help="Window type for FFT (default: hann)"
    )
    
    parser.add_argument(
        "--export-binned-jpgs",
        action="store_true",
        help="Export per-chunk JPG visualizations for binned analysis (mean power, percent power, dominant band per chunk; occupancy once)"
    )
    
    parser.add_argument(
        "--export-binned-csvs",
        action="store_true",
        help="Export CSV/Excel summaries for binned analysis"
    )

    parser.add_argument(
        "--plot-trajectory",
        action="store_true",
        help="Export trajectory plot with EOI overlays (if available)"
    )

    parser.add_argument(
        "--eoi-file",
        type=str,
        default=None,
        help="Optional EOI file or directory to use instead of auto-detect"
    )
    
    args = parser.parse_args()
    
    # Get speed thresholds
    low_speed = args.speed_min
    high_speed = args.speed_max
    
    # Validate speed range
    if low_speed < 0 or high_speed < 0:
        print(f"✗ Error: Speed thresholds must be non-negative")
        sys.exit(1)
    if low_speed >= high_speed:
        print(f"✗ Error: --speed-min ({low_speed}) must be less than --speed-max ({high_speed})")
        sys.exit(1)
    
    # Check if input is a directory or file
    if os.path.isdir(args.input_path):
        # Directory mode - process all recordings
        print(f"\n{'='*60}")
        print("Spatial Spectral Mapper - Batch Directory Processing")
        print(f"{'='*60}")
        print(f"Directory: {args.input_path}")
        print(f"Parameters: PPM={args.ppm}, Chunk={args.chunk_size}s, Speed={low_speed}-{high_speed} cm/s")
        print(f"{'='*60}\n")
        
        recordings = find_recording_files(args.input_path)
        
        if not recordings:
            print("✗ No valid recording files found in directory.")
            sys.exit(1)
        
        print(f"Found {len(recordings)} recording(s) to process\n")
        
        success_count = 0
        fail_count = 0
        
        for idx, (electrophys_file, pos_file) in enumerate(recordings, 1):
            print(f"\n[{idx}/{len(recordings)}] Processing: {os.path.basename(electrophys_file)}")
            print("-" * 60)
            
            try:
                process_single_file(
                    electrophys_file,
                    pos_file,
                    args.output or os.path.dirname(electrophys_file),
                    args.ppm,
                    args.chunk_size,
                    low_speed,
                    high_speed,
                    args.window,
                    export_binned_jpgs=args.export_binned_jpgs,
                    export_binned_csvs=args.export_binned_csvs,
                    plot_trajectory=args.plot_trajectory,
                    eoi_file_override=args.eoi_file
                )
                success_count += 1
            except Exception as e:
                print(f"✗ Failed: {e}")
                fail_count += 1
        
        print(f"\n{'='*60}")
        print(f"Batch processing complete!")
        print(f"✓ Successful: {success_count}")
        if fail_count > 0:
            print(f"✗ Failed: {fail_count}")
        print(f"{'='*60}")
        
    else:
        # Single file mode
        if not os.path.exists(args.input_path):
            print(f"✗ Error: File not found: {args.input_path}")
            sys.exit(1)
        
        try:
            pos_file = find_pos_file(args.input_path)
        except FileNotFoundError as e:
            print(f"✗ Error: {e}")
            sys.exit(1)
        
        output_dir = args.output or os.path.dirname(args.input_path) or "."
        
        print(f"\n{'='*60}")
        print("Spatial Spectral Mapper - Single File Processing")
        print(f"{'='*60}")
        
        try:
            process_single_file(
                args.input_path, 
                pos_file, 
                output_dir,
                args.ppm,
                args.chunk_size,
                low_speed,
                high_speed,
                args.window,
                    export_binned_jpgs=args.export_binned_jpgs,
                export_binned_csvs=args.export_binned_csvs,
                plot_trajectory=args.plot_trajectory,
                eoi_file_override=args.eoi_file
            )
        except Exception as e:
            print(f"\n✗ Error during processing: {e}")
            import traceback
            traceback.print_exc()
            sys.exit(1)


class TimeoutError(Exception):
    pass

def timeout_handler(signum, frame):
    raise TimeoutError("Processing timeout exceeded")

def process_single_file(electrophys_file, pos_file, output_dir, ppm, chunk_size, 
                        low_speed, high_speed, window_type, export_binned_jpgs=False, export_binned_csvs=False, plot_trajectory=False, eoi_file_override=None, timeout_seconds=300):
    """Process a single EEG/EGF file with timeout protection"""
    
    # Monitor memory at start
    process = psutil.Process()
    mem_before = process.memory_info().rss / 1024 / 1024  # MB
    
    # Generate output filename
    base_name = os.path.splitext(os.path.basename(electrophys_file))[0]
    output_csv = os.path.join(output_dir, f"{base_name}_SSM.csv")
    
    print(f"Input:  {electrophys_file}")
    print(f"Output: {output_csv}")
    print(f"PPM: {ppm}, Chunk: {chunk_size}s, Speed: {low_speed}-{high_speed} cm/s")
    print(f"Memory: {mem_before:.1f} MB")
    
    
    # Create batch worker with mock signals
    worker = BatchWorker(output_dir)
    
    print("[1/5] Starting data processing...")
    print(f"  Timeout set to {timeout_seconds} seconds")

    try:
        # Run initialization with worker as 'self' parameter
        result = initialize_fMap(
            worker,
            files=[pos_file, electrophys_file],
            ppm=ppm,
            chunk_size=chunk_size,
            window_type=window_type,
            low_speed=low_speed,
            high_speed=high_speed
        )
        print("[2/5] Data processing complete")
    except Exception as e:
        print(f"\n✗ Error during data processing: {e}")
        import traceback
        traceback.print_exc()
        raise

    
    
    # Unpack legacy/new result signature
    arena_shape = None
    if len(result) == 6:
        freq_maps, plot_data, pos_t, scaling_factor_crossband, chunk_pows_data, tracking_data = result
        binned_data = None
    elif len(result) == 7:
        freq_maps, plot_data, pos_t, scaling_factor_crossband, chunk_pows_data, tracking_data, binned_data = result
    else:
        freq_maps, plot_data, pos_t, scaling_factor_crossband, chunk_pows_data, tracking_data, binned_data, arena_shape = result
    
    print("[3/5] Extracting tracking data...")
    # Extract tracking data
    if tracking_data:
        # Handle 3-tuple (x, y, t) or 2-tuple (x, y)
        pos_x_chunks = tracking_data[0]
        pos_y_chunks = tracking_data[1]
    else:
        pos_x_chunks, pos_y_chunks = None, None
    
    # Process EOIs if available
    eoi_segments = {}
    eoi_file = resolve_eoi_file(electrophys_file, eoi_file_override)
    if eoi_file_override and not eoi_file:
        print(f"  ⚠ Explicit EOI file not found: {eoi_file_override}")
    if eoi_file and tracking_data:
        print(f"[3.5/5] Processing EOIs from {os.path.basename(eoi_file)}...")
        eois = load_eois(eoi_file)
        if eois:
            # Map EOIs to spatial segments
            # tracking_data is (x_chunks, y_chunks, t_chunks)
            if len(tracking_data) == 3:
                x_chunks, y_chunks, t_chunks = tracking_data
                for i, (chunk_x, chunk_y, chunk_t) in enumerate(zip(x_chunks, y_chunks, t_chunks)):
                    if len(chunk_t) == 0: continue
                    for eoi in eois:
                        start = eoi["start"]
                        stop = eoi["stop"]
                        label = eoi.get("label")
                        # Find indices where time is within EOI
                        # Use searchsorted for speed
                        idx_start = np.searchsorted(chunk_t, start)
                        idx_end = np.searchsorted(chunk_t, stop)
                        
                        if idx_end > idx_start:
                            if i not in eoi_segments: eoi_segments[i] = []
                            eoi_segments[i].append((chunk_x[idx_start:idx_end], chunk_y[idx_start:idx_end], label))
            print(f"  → Mapped {len(eois)} EOIs to {sum(len(v) for v in eoi_segments.values())} spatial segments")
    elif eoi_file:
        print(f"  ⚠ EOI file found but no tracking data available.")

    if plot_trajectory and tracking_data:
        trajectory_path = os.path.join(output_dir, f"{base_name}_trajectory_eoi.jpg")
        print("[3.8/5] Exporting trajectory plot...")
        plot_trajectory_with_eois(
            trajectory_path,
            pos_x_chunks,
            pos_y_chunks,
            eoi_segments,
            ppm=ppm,
            arena_shape=arena_shape,
            binned_data=binned_data
        )
        print(f"  → Trajectory plot saved to: {trajectory_path}")

    
    print("[4/5] Exporting to CSV...")
    # Export to CSV
    export_to_csv(
        output_csv,
        pos_t,
        chunk_pows_data,
        chunk_size,
        ppm,
        pos_x_chunks,
        pos_y_chunks
    )
    print("[5/5] SSM data export complete")

    # Export binned analysis outputs if available
    try:
        if binned_data is None:
            pass
        elif not (export_binned_jpgs or export_binned_csvs):
            print("  → Binned data available but neither --export-binned-jpgs nor --export-binned-csvs set; skipping binned export.")
        else:
            output_folder = os.path.join(output_dir, f"{base_name}_binned_analysis")
            os.makedirs(output_folder, exist_ok=True)
            
            if binned_data.get('type') == 'polar':
                # Export polar binned analysis in batch: produce JPG visualizations and CSV summaries
                print("  → Polar binned analysis detected.")

                bands = binned_data['bands']
                n_chunks = binned_data['time_chunks']

                # Export per-chunk polar power and percent-power (2x8 grids per band)
                from matplotlib.ticker import FuncFormatter
                theta = np.linspace(-np.pi, np.pi, 9)
                r = [0, 1.0/np.sqrt(2), 1]
                T, R = np.meshgrid(theta, r)

                if export_binned_jpgs:
                    print("  → Exporting polar JPGs...")
                    export_count = 0

                    # Setup grid for polar pcolormesh
                    theta = np.linspace(-np.pi, np.pi, 9)
                    r = [0, 1.0/np.sqrt(2), 1]
                    T, R = np.meshgrid(theta, r)

                    # Precompute total power per chunk for percent calculations
                    for chunk_idx in range(n_chunks):
                        # Mean power (per chunk)
                        fig, axes = plt.subplots(2, 4, figsize=(16, 8), subplot_kw={'projection': 'polar'})
                        fig.suptitle(f'Polar Bins - Chunk {chunk_idx + 1:02d} (Frequency Band Power & Occupancy)', fontsize=14, fontweight='bold')
                        ax_flat = axes.flatten()

                        for idx, band in enumerate(bands):
                            if idx >= len(ax_flat):
                                break
                            ax = ax_flat[idx]
                            data = binned_data['bin_power_timeseries'][band][:, :, chunk_idx]
                            im = ax.pcolormesh(T, R, data, cmap=BATCH_CMAP_POWER, shading='flat')
                            ax.set_title(band)
                            ax.set_yticklabels([])
                            ax.set_xticklabels([])
                            ax.grid(True, alpha=0.3)
                            cbar = plt.colorbar(im, ax=ax, pad=0.05, shrink=0.8)
                            cbar.set_label('Power', fontsize=8)

                        # Occupancy slot (8th)
                        if len(bands) < 8:
                            ax_occ = ax_flat[7]
                            occ_data = None
                            if 'bin_occupancy_timeseries' in binned_data:
                                occ_data = binned_data['bin_occupancy_timeseries'][:, :, chunk_idx]
                                total = np.sum(occ_data)
                                if total > 0:
                                    occ_data = (occ_data / total) * 100.0
                            else:
                                occ = binned_data.get('bin_occupancy')
                                if occ is not None:
                                    total = np.sum(occ)
                                    if total > 0:
                                        occ_data = (occ / total) * 100.0

                            if occ_data is not None:
                                im = ax_occ.pcolormesh(T, R, occ_data, cmap=BATCH_CMAP_OCCUPANCY, shading='flat', vmin=0, vmax=100)
                                ax_occ.set_title('Occupancy (%)')
                                ax_occ.set_yticklabels([])
                                ax_occ.set_xticklabels([])
                                ax_occ.grid(True, alpha=0.3)
                                cbar = plt.colorbar(im, ax=ax_occ, pad=0.05, shrink=0.8)
                                cbar.set_label('%', fontsize=8)
                            else:
                                ax_occ.axis('off')

                        for idx in range(len(bands), 8):
                            ax_flat[idx].axis('off')

                        plt.tight_layout()
                        jpg_path = os.path.join(output_folder, f"{base_name}_chunk{chunk_idx+1:02d}_polar_power.jpg")
                        fig.savefig(jpg_path, format='jpg', pil_kwargs={'quality':85}, bbox_inches='tight')
                        plt.close(fig)
                        export_count += 1

                        # Percent power (per chunk)
                        fig, axes = plt.subplots(2, 4, figsize=(16, 8), subplot_kw={'projection': 'polar'})
                        fig.suptitle(f'Polar Bins - Chunk {chunk_idx + 1:02d} (Frequency Band Percent Power & Occupancy)', fontsize=14, fontweight='bold')
                        ax_flat = axes.flatten()

                        total = np.zeros((2, 8))
                        for band in bands:
                            total += binned_data['bin_power_timeseries'][band][:, :, chunk_idx]

                        for idx, band in enumerate(bands):
                            if idx >= len(ax_flat):
                                break
                            ax = ax_flat[idx]
                            band_power = binned_data['bin_power_timeseries'][band][:, :, chunk_idx]
                            with np.errstate(divide='ignore', invalid='ignore'):
                                pct = np.where(total > 0, (band_power / total) * 100.0, 0.0)
                            im = ax.pcolormesh(T, R, pct, cmap=BATCH_CMAP_PERCENT, shading='flat', vmin=0, vmax=100)
                            ax.set_title(band)
                            ax.set_yticklabels([])
                            ax.set_xticklabels([])
                            ax.grid(True, alpha=0.3)
                            cbar = plt.colorbar(im, ax=ax, pad=0.05, shrink=0.8)
                            cbar.set_label('%', fontsize=8)

                        # Occupancy slot (8th)
                        if len(bands) < 8:
                            ax_occ = ax_flat[7]
                            occ_data = None
                            if 'bin_occupancy_timeseries' in binned_data:
                                occ_data = binned_data['bin_occupancy_timeseries'][:, :, chunk_idx]
                                total = np.sum(occ_data)
                                if total > 0:
                                    occ_data = (occ_data / total) * 100.0
                            else:
                                occ = binned_data.get('bin_occupancy')
                                if occ is not None:
                                    total = np.sum(occ)
                                    if total > 0:
                                        occ_data = (occ / total) * 100.0

                            if occ_data is not None:
                                im = ax_occ.pcolormesh(T, R, occ_data, cmap=BATCH_CMAP_OCCUPANCY, shading='flat', vmin=0, vmax=100)
                                ax_occ.set_title('Occupancy (%)')
                                ax_occ.set_yticklabels([])
                                ax_occ.set_xticklabels([])
                                ax_occ.grid(True, alpha=0.3)
                                cbar = plt.colorbar(im, ax=ax_occ, pad=0.05, shrink=0.8)
                                cbar.set_label('%', fontsize=8)
                            else:
                                ax_occ.axis('off')

                        for idx in range(len(bands), 7):
                            ax_flat[idx].axis('off')

                        plt.tight_layout()
                        jpg_path = os.path.join(output_folder, f"{base_name}_chunk{chunk_idx+1:02d}_polar_percent.jpg")
                        fig.savefig(jpg_path, format='jpg', pil_kwargs={'quality':85}, bbox_inches='tight')
                        plt.close(fig)
                        export_count += 1

                        # Dominant + EOI combined (1x2 polar)
                        fig, axes = plt.subplots(1, 2, figsize=(12, 5), subplot_kw={'projection': 'polar'})
                        theta = np.linspace(-np.pi, np.pi, 9)
                        r = [0, 1.0/np.sqrt(2), 1]
                        T, R = np.meshgrid(theta, r)

                        # Dominant
                        dominant_chunk = binned_data['bin_dominant_band'][chunk_idx]
                        band_map = {band: idx for idx, band in enumerate(bands)}
                        numeric_dominant = np.zeros((2, 8))
                        for ri in range(2):
                            for si in range(8):
                                numeric_dominant[ri, si] = band_map.get(dominant_chunk[ri, si], 0)
                        im1 = axes[0].pcolormesh(T, R, numeric_dominant, cmap=BATCH_CMAP_DOMINANT, shading='flat', vmin=0, vmax=len(bands)-1)
                        axes[0].set_title(f'Dominant Band - Chunk {chunk_idx+1:02d}')
                        axes[0].set_yticklabels([])
                        cbar1 = plt.colorbar(im1, ax=axes[0], ticks=range(len(bands)), pad=0.1)
                        cbar1.set_ticklabels(bands, fontsize=8)

                        # EOI distribution
                        H = None
                        if eoi_segments and chunk_idx in eoi_segments:
                            H = np.zeros((2, 8))
                            for seg in eoi_segments[chunk_idx]:
                                xs = np.array(seg[0])
                                ys = np.array(seg[1])
                                if xs.size == 0:
                                    continue
                                # map positions into polar bins using equal-area transform
                                all_x = np.concatenate(pos_x_chunks) if pos_x_chunks else np.array([])
                                all_y = np.concatenate(pos_y_chunks) if pos_y_chunks else np.array([])
                                if all_x.size and all_y.size:
                                    min_x, max_x = np.min(all_x), np.max(all_x)
                                    min_y, max_y = np.min(all_y), np.max(all_y)
                                else:
                                    min_x, max_x, min_y, max_y = 0, 1, 0, 1
                                width = max_x - min_x if max_x > min_x else 1
                                height = max_y - min_y if max_y > min_y else 1
                                nx = 2 * (xs - min_x) / width - 1
                                ny = 2 * (ys - min_y) / height - 1
                                r_pt = np.sqrt(nx**2 + ny**2)
                                theta_pt = np.arctan2(ny, nx)
                                equal_area_radius = 1.0 / np.sqrt(2.0)
                                r_bins = [0, equal_area_radius, np.inf]
                                r_indices = np.clip(np.digitize(r_pt, r_bins) - 1, 0, 1)
                                theta_edges = np.linspace(-np.pi, np.pi, 9)
                                th_indices = np.clip(np.digitize(theta_pt, theta_edges) - 1, 0, 7)
                                for ri, ti in zip(r_indices, th_indices):
                                    H[ri, ti] += 1

                        if H is None:
                            axes[1].axis('off')
                        else:
                            im2 = axes[1].pcolormesh(T, R, H, cmap=BATCH_CMAP_OCCUPANCY, shading='flat')
                            axes[1].set_title(f'EOI Distribution - Chunk {chunk_idx+1:02d}')
                            axes[1].set_yticklabels([])
                            cbar2 = plt.colorbar(im2, ax=axes[1], pad=0.1)
                            cbar2.set_label('EOI Count', fontsize=8)

                        plt.tight_layout()
                        jpg_path = os.path.join(output_folder, f"{base_name}_chunk{chunk_idx+1:02d}_polar_dominant_eoi.jpg")
                        fig.savefig(jpg_path, format='jpg', pil_kwargs={'quality':85}, bbox_inches='tight')
                        plt.close(fig)
                        export_count += 1

                    print(f"  ✓ Exported {export_count} polar JPG(s)")

                if export_binned_csvs:
                    print("  → Exporting polar data (CSV/Excel)...")
                    output_prefix = os.path.join(output_folder, f"{base_name}_binned")
                    
                    # Helper to reshape (2, 8, n_chunks) -> (n_chunks, 16) and add Chunk column
                    def prepare_matrix(data_3d):
                        flat = data_3d.reshape(16, -1).T
                        chunks_col = np.arange(1, n_chunks + 1).reshape(-1, 1)
                        return np.hstack([chunks_col, flat])

                    # 1. Mean Power (per chunk)
                    power_per_chunk = {}
                    for band in bands:
                        power_per_chunk[band] = prepare_matrix(binned_data['bin_power_timeseries'][band])

                    # 2. Percent Power (per chunk)
                    percent_per_chunk = {}
                    all_bands_power = np.stack([binned_data['bin_power_timeseries'][b] for b in bands])
                    total_power = np.sum(all_bands_power, axis=0) # (2, 8, n_chunks)
                    for band in bands:
                        with np.errstate(divide='ignore', invalid='ignore'):
                            pct_data = np.where(total_power > 0, (binned_data['bin_power_timeseries'][band] / total_power) * 100.0, 0.0)
                        percent_per_chunk[band] = prepare_matrix(pct_data)

                    # 3. EOIs (per chunk)
                    eoi_export_matrix = None
                    if eoi_segments:
                        eoi_counts_per_chunk = np.zeros((2, 8, n_chunks))
                        all_x = np.concatenate(pos_x_chunks)
                        all_y = np.concatenate(pos_y_chunks)
                        min_x, max_x = np.min(all_x), np.max(all_x)
                        min_y, max_y = np.min(all_y), np.max(all_y)
                        width = max_x - min_x
                        height = max_y - min_y
                        if width == 0: width = 1
                        if height == 0: height = 1

                        for chunk_idx in range(n_chunks):
                            if chunk_idx in eoi_segments:
                                for seg in eoi_segments[chunk_idx]:
                                    nx = 2 * (np.array(seg[0]) - min_x) / width - 1
                                    ny = 2 * (np.array(seg[1]) - min_y) / height - 1
                                    r = np.sqrt(nx**2 + ny**2)
                                    theta = np.arctan2(ny, nx)
                                    
                                    equal_area_radius = 1.0 / np.sqrt(2.0)
                                    r_bins = [0, equal_area_radius, np.inf]
                                    r_indices = np.clip(np.digitize(r, r_bins) - 1, 0, 1)
                                    
                                    theta_edges = np.linspace(-np.pi, np.pi, 9)
                                    th_indices = np.clip(np.digitize(theta, theta_edges) - 1, 0, 7)
                                    
                                    for ri, ti in zip(r_indices, th_indices):
                                        eoi_counts_per_chunk[ri, ti, chunk_idx] += 1
                        eoi_export_matrix = prepare_matrix(eoi_counts_per_chunk)

                    # 4. Percent Occupancy (per chunk)
                    occupancy_export_matrix = None
                    if 'bin_occupancy_timeseries' in binned_data:
                        occ_ts = binned_data['bin_occupancy_timeseries']
                        occ_sums = np.sum(occ_ts, axis=(0, 1))
                        with np.errstate(divide='ignore', invalid='ignore'):
                            occ_pct = np.where(occ_sums[None, None, :] > 0, (occ_ts / occ_sums[None, None, :]) * 100.0, 0.0)
                        occupancy_export_matrix = prepare_matrix(occ_pct)

                    # 5. Dominant Band (per chunk)
                    dom_data = np.array(binned_data['bin_dominant_band']) # (n_chunks, 2, 8)
                    dom_data_T = dom_data.transpose(1, 2, 0) # (2, 8, n_chunks)
                    dom_export_matrix = prepare_matrix(dom_data_T)

                    try:
                        import openpyxl
                        bin_headers = ["Chunk"] + [f"Inner_S{i+1}" for i in range(8)] + [f"Outer_S{i+1}" for i in range(8)]
                        
                        # Helper to save multi-sheet workbook
                        def save_multi_sheet(data_dict, filename):
                            wb = openpyxl.Workbook()
                            wb.remove(wb.active)
                            for sheet_name, matrix in data_dict.items():
                                ws = wb.create_sheet(title=sheet_name)
                                ws.append(bin_headers)
                                for row in matrix:
                                    ws.append([float(val) for val in row])
                            wb.save(filename)

                        # Helper to save single-sheet workbook
                        def save_single_sheet(matrix, filename, sheet_title, is_string=False):
                            wb = openpyxl.Workbook()
                            ws = wb.active
                            ws.title = sheet_title
                            ws.append(bin_headers)
                            for row in matrix:
                                if is_string:
                                    ws.append([row[0]] + list(row[1:]))
                                else:
                                    ws.append([float(val) for val in row])
                            wb.save(filename)

                        save_multi_sheet(power_per_chunk, f"{output_prefix}_mean_power_per_chunk.xlsx")
                        save_multi_sheet(percent_per_chunk, f"{output_prefix}_percent_power_per_chunk.xlsx")
                        
                        if eoi_export_matrix is not None:
                            save_single_sheet(eoi_export_matrix, f"{output_prefix}_eois_per_chunk.xlsx", "EOIs")
                        
                        if occupancy_export_matrix is not None:
                            save_single_sheet(occupancy_export_matrix, f"{output_prefix}_percent_occupancy_per_chunk.xlsx", "Percent Occupancy")
                            
                        save_single_sheet(dom_export_matrix, f"{output_prefix}_dominant_band_per_chunk.xlsx", "Dominant Band", is_string=True)

                        print(f"  ✓ Excel files exported to: {output_folder}")
                        
                    except ImportError:
                        # CSV Fallback
                        for band in bands:
                            np.savetxt(f"{output_prefix}_mean_power_{band}.csv", power_per_chunk[band], delimiter=',')
                            np.savetxt(f"{output_prefix}_percent_power_{band}.csv", percent_per_chunk[band], delimiter=',')
                        if eoi_export_matrix is not None:
                            np.savetxt(f"{output_prefix}_eois.csv", eoi_export_matrix, delimiter=',')
                        if occupancy_export_matrix is not None:
                            np.savetxt(f"{output_prefix}_percent_occupancy.csv", occupancy_export_matrix, delimiter=',')
                        # Dominant band CSV might be tricky with strings if using np.savetxt with default fmt
                        np.savetxt(f"{output_prefix}_dominant_band.csv", dom_export_matrix, delimiter=',', fmt='%s')
                        print(f"  ✓ CSVs exported to: {output_folder}")

            else:
                # Rectangular (4x4)
                output_prefix = os.path.join(output_folder, base_name + "_binned")
                
                if export_binned_csvs:
                    print("  → Exporting binned data (CSV/Excel)...")
                    result = export_binned_analysis_to_csv(binned_data, output_prefix, pos_x_chunks=pos_x_chunks, pos_y_chunks=pos_y_chunks, eoi_segments=eoi_segments)
                    if result.get('format') == 'csv' and result.get('reason') == 'openpyxl_not_installed':
                        print("  ⚠ Excel export unavailable (openpyxl missing). Exported CSVs.")
                    else:
                        print(f"  ✓ Exported {result.get('format').upper()} files")
                
                if export_binned_jpgs:
                    print("  → Exporting binned JPGs...")
                    try:
                        jpg_count = export_binned_analysis_jpgs(binned_data, output_folder, base_name)
                        print(f"  ✓ Exported {jpg_count} JPGs")
                    except Exception as e:
                        print(f"  ⚠ Failed to export JPGs: {e}")
    except Exception as e:
        print(f"⚠ Binned analysis export failed: {e}")
    
    # Print summary
    print(f"✓ Complete - {len(pos_t)} time bins processed")
    for band, power in scaling_factor_crossband.items():
        print(f"  {band:15s}: {power*100:6.2f}%")
    
    # Explicitly delete large objects to free memory
    del freq_maps, plot_data, pos_t, chunk_pows_data, tracking_data
    del pos_x_chunks, pos_y_chunks, worker
    
    # Force garbage collection to clear memory before next file
    gc.collect()
    
    # Monitor memory after cleanup
    mem_after = process.memory_info().rss / 1024 / 1024  # MB
    print(f"Memory after cleanup: {mem_after:.1f} MB (freed {mem_before - mem_after:.1f} MB)")


if __name__ == "__main__":
    main()
