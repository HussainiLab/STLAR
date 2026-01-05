# -*- coding: utf-8 -*-
"""
Created on Thu May 27 12:51:19 2021
Updated on Dec 21 2025
@author: vajramsrujan
@author: Hussainilab
"""

import os
import sys
import matplotlib
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
import numpy as np
import csv
import glob
from pathlib import Path
import subprocess
from matplotlib.patches import Ellipse

from PyQt5.QtWidgets import QMessageBox
from functools import partial
from initialize_fMap import initialize_fMap
from matplotlib import cm
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg
from matplotlib.figure import Figure
from core.data_loaders import grab_position_data
from core.processors.Tint_Matlab import speed2D
from core.processors.spectral_functions import (
    speed_bins,
    export_binned_analysis_to_csv,
    visualize_binned_analysis,
    visualize_binned_analysis_by_chunk,
    visualize_binned_occupancy_and_dominant
)
from core.worker_thread import WorkerSignals
from PyQt5.QtGui import QPixmap, QFont, QImage
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer
from PyQt5.QtWidgets import *
from core.worker_thread.Worker import Worker

matplotlib.use('Qt5Agg')

# GUI colormap defaults (GUI path only; batch uses spectral_functions constants)
GUI_CMAP_POWER = 'turbo'
GUI_CMAP_PERCENT = 'turbo'
GUI_CMAP_OCCUPANCY = 'turbo'
GUI_CMAP_DOMINANT = 'tab10'
    
# =========================================================================== #

class BatchSignals(QThread):
    '''Wrapper class to provide signals interface compatible with initialize_fMap'''
    def __init__(self):
        self.progress = None
        self.text_progress = None

class BatchProgressSignal:
    '''Signal wrapper for progress updates'''
    def __init__(self, parent_signal):
        self.parent_signal = parent_signal
    
    def emit(self, value):
        if self.parent_signal:
            self.parent_signal.emit(value)

class BatchTextSignal:
    '''Signal wrapper for text progress updates'''
    def __init__(self, parent_signal):
        self.parent_signal = parent_signal
    
    def emit(self, value):
        if self.parent_signal:
            self.parent_signal.emit(value)

class BatchWorkerThread(QThread):
    '''
        Worker thread for batch processing multiple files in a folder.
        Emits signals for progress updates.
    '''
    progress_update = pyqtSignal(str)  # For text progress updates
    progress_value = pyqtSignal(int)   # For progress bar
    finished_signal = pyqtSignal(dict) # For completion with results
    error_signal = pyqtSignal(str)     # For error messages
    
    def __init__(self, folder_path, ppm, chunk_size, window_type, 
                 low_speed, high_speed, output_dir=None):
        super().__init__()
        self.folder_path = folder_path
        self.ppm = ppm
        self.chunk_size = chunk_size
        self.window_type = window_type
        self.low_speed = low_speed
        self.high_speed = high_speed
        self.output_dir = output_dir or folder_path
        self.total_files = 0
        self.processed_files = 0
        self.successful_files = []
        self.failed_files = []
        
        # Create signals wrapper for compatibility with initialize_fMap
        self.signals = BatchSignals()
        self.signals.progress = BatchProgressSignal(self.progress_value)
        self.signals.text_progress = BatchTextSignal(self.progress_update)
    
    def find_recording_files(self):
        '''
            Find electrophysiology recordings in folder by priority:
              1) .egf
              2) .egf2, .egf3, .egf4
              3) .eeg
              4) .eeg2, .eeg3, .eeg4
            Returns: (recordings_list, missing_pos_bases, found_any_ephys)
        '''
        recordings = {}
        
        # Gather all candidate files across prioritized extensions
        patterns = ["*.egf", "*.egf2", "*.egf3", "*.egf4",
                    "*.eeg", "*.eeg2", "*.eeg3", "*.eeg4"]
        all_files = []
        for pat in patterns:
            all_files.extend(glob.glob(os.path.join(self.folder_path, pat)))
        
        # Group by basename (without extension)
        base_to_files = {}
        for fpath in all_files:
            base = Path(fpath).stem
            ext = Path(fpath).suffix.lower().lstrip('.')  # e.g., 'egf', 'egf2', 'eeg3'
            base_to_files.setdefault(base, {})[ext] = fpath
        
        # Selection priority
        priority = ["egf", "egf2", "egf3", "egf4",
                    "eeg", "eeg2", "eeg3", "eeg4"]
        
        missing_pos_bases = []

        # For each base, pick best available by priority and ensure .pos exists
        for base, ext_map in base_to_files.items():
            chosen = None
            for ext in priority:
                if ext in ext_map:
                    chosen = ext_map[ext]
                    break
            if not chosen:
                continue
            pos_file = self._find_pos_file(chosen)
            if pos_file:
                recordings[base] = (chosen, pos_file)
            else:
                missing_pos_bases.append(base)
        
        found_any_ephys = len(base_to_files) > 0
        return list(recordings.values()), missing_pos_bases, found_any_ephys
    
    def _find_pos_file(self, eeg_file):
        '''Auto-detect .pos file based on .eeg/.egf filename'''
        base_name = os.path.splitext(eeg_file)[0]
        pos_file = base_name + '.pos'
        
        if os.path.exists(pos_file):
            return pos_file
        return None
    
    def run(self):
        '''Execute batch processing in worker thread'''
        try:
            recordings, missing_pos_bases, found_any_ephys = self.find_recording_files()
            self.total_files = len(recordings)
            
            # Report missing .pos per base
            if missing_pos_bases:
                bases_preview = ", ".join(missing_pos_bases[:5])
                more_count = len(missing_pos_bases) - 5
                suffix = f" ... and {more_count} more" if more_count > 0 else ""
                self.error_signal.emit(
                    f"Missing .pos files for {len(missing_pos_bases)} base(s): {bases_preview}{suffix}"
                )

            # If no recordings and no ephys found at all, show specific message
            if not recordings and not found_any_ephys:
                self.error_signal.emit(
                    f"No electrophysiology files found in {self.folder_path}. Expected .egf/.egf2-4 or .eeg/.eeg2-4."
                )
                self.finished_signal.emit({'successful': [], 'failed': []})
                return
            
            # If some ephys were found but none had .pos, end gracefully
            if not recordings and found_any_ephys:
                self.error_signal.emit(
                    f"No recordings could be processed because matching .pos files were missing."
                )
                self.finished_signal.emit({'successful': [], 'failed': []})
                return
            
            self.progress_update.emit(f"Found {self.total_files} file(s) to process")
            
            for electrophys_file, pos_file in recordings:
                try:
                    self.processed_files += 1
                    filename = os.path.basename(electrophys_file)
                    
                    self.progress_update.emit(f"[{self.processed_files}/{self.total_files}] Processing: {filename}")
                    self.progress_value.emit(int((self.processed_files / self.total_files) * 100))
                    
                    # Call initialize_fMap to process the file
                    result = initialize_fMap(
                        self,  # Pass self as the worker/signals object
                        files=[pos_file, electrophys_file],
                        ppm=self.ppm,
                        chunk_size=self.chunk_size,
                        window_type=self.window_type,
                        low_speed=self.low_speed,
                        high_speed=self.high_speed
                    )
                    
                    # Export results to CSV
                    self._export_to_csv(result, electrophys_file)
                    self.successful_files.append(filename)
                    self.progress_update.emit(f"  ✓ Complete: {filename}")
                    
                except Exception as e:
                    self.failed_files.append((os.path.basename(electrophys_file), str(e)))
                    self.progress_update.emit(f"  ✗ Failed: {os.path.basename(electrophys_file)} - {str(e)}")
            
            # Emit completion signal
            results = {
                'successful': self.successful_files,
                'failed': self.failed_files,
                'total': self.total_files
            }
            self.finished_signal.emit(results)
            
        except Exception as e:
            self.error_signal.emit(f"Batch processing error: {str(e)}")
            self.finished_signal.emit({'successful': [], 'failed': []})
    
    def _export_to_csv(self, result, electrophys_file):
        '''Export processing results to CSV and binned analysis'''
        # Support legacy (6-tuple) and new (7-tuple with binned_data)
        if len(result) == 6:
            freq_maps, plot_data, pos_t, scaling_factor_crossband, chunk_powers_data, tracking_data = result
            binned_data = None
        elif len(result) == 7:
            freq_maps, plot_data, pos_t, scaling_factor_crossband, chunk_powers_data, tracking_data, binned_data = result
        else:
            freq_maps, plot_data, pos_t, scaling_factor_crossband, chunk_powers_data, tracking_data, binned_data, arena_shape = result
        
        base_name = os.path.splitext(os.path.basename(electrophys_file))[0]
        output_csv = os.path.join(self.output_dir, f"{base_name}_SSM.csv")
        
        bands = [
            ("Delta", "Avg Delta Power"),
            ("Theta", "Avg Theta Power"),
            ("Beta", "Avg Beta Power"),
            ("Low Gamma", "Avg Low Gamma Power"),
            ("High Gamma", "Avg High Gamma Power"),
            ("Ripple", "Avg Ripple Power"),
            ("Fast Ripple", "Avg Fast Ripple Power"),
        ]
        
        band_labels = [label for _, label in bands]
        percent_labels = [f"Percent {name}" for name, _ in bands]
        
        # Calculate distances for each chunk
        distances_per_bin = []
        cumulative_distances = []
        cumulative_sum = 0.0
        
        if tracking_data:
            if len(tracking_data) == 3:
                pos_x_chunks, pos_y_chunks, _ = tracking_data
            else:
                pos_x_chunks, pos_y_chunks = tracking_data
            
            for i in range(len(pos_x_chunks)):
                distance_cm_in_bin = 0.0
                
                x_bin = pos_x_chunks[i]
                y_bin = pos_y_chunks[i]
                
                if len(x_bin) > 1:
                    dx = np.diff(np.array(x_bin))
                    dy = np.diff(np.array(y_bin))
                    distances_in_bin_pixels = np.sqrt(dx**2 + dy**2)
                    total_distance_pixels_in_bin = np.sum(distances_in_bin_pixels)
                    
                    if self.ppm is not None and self.ppm > 0:
                        distance_cm_in_bin = (total_distance_pixels_in_bin / self.ppm) * 100
                    else:
                        distance_cm_in_bin = total_distance_pixels_in_bin
                
                distances_per_bin.append(distance_cm_in_bin)
                cumulative_sum += distance_cm_in_bin
                cumulative_distances.append(cumulative_sum)
        
        # Write CSV
        header = ["Time Bin (s)", "Distance Per Bin (cm)", "Cumulative Distance (cm)"] + band_labels + percent_labels
        
        with open(output_csv, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(header)
            
            actual_duration = float(pos_t[-1])
            max_full_chunks = int(actual_duration / self.chunk_size)
            num_rows = min(len(pos_t), max_full_chunks)
            
            # Gather per-band arrays
            band_arrays = {}
            for key, label in bands:
                arr = np.array(chunk_powers_data.get(key, [])).reshape(-1)
                band_arrays[label] = arr
            
            for i in range(num_rows):
                time_bin_start = i * self.chunk_size
                time_bin_end = (i + 1) * self.chunk_size
                time_bin_str = f"{time_bin_start}-{time_bin_end}"
                
                row = [time_bin_str]
                
                if distances_per_bin and i < len(distances_per_bin):
                    row.append(round(distances_per_bin[i], 3))
                else:
                    row.append("")
                
                if cumulative_distances and i < len(cumulative_distances):
                    row.append(round(cumulative_distances[i], 3))
                else:
                    row.append("")
                
                band_values = []
                for _, label in bands:
                    val = band_arrays[label][i] if i < len(band_arrays[label]) else ""
                    band_values.append(val)
                
                row.extend(band_values)
                
                numeric_vals = [float(v) for v in band_values if v != ""]
                total_power = sum(numeric_vals)
                
                for v in band_values:
                    if v == "" or total_power == 0:
                        row.append("")
                    else:
                        row.append(round((float(v) / total_power) * 100.0, 3))
                
                writer.writerow(row)

        # Export binned analysis (Excel + JPG visualization) if available
        try:
            if binned_data is not None:
                output_prefix = os.path.join(self.output_dir, f"{base_name}_binned")
                pos_x_chunks = tracking_data[0] if tracking_data else None
                pos_y_chunks = tracking_data[1] if tracking_data else None
                export_result = export_binned_analysis_to_csv(binned_data, output_prefix, pos_x_chunks=pos_x_chunks, pos_y_chunks=pos_y_chunks, eoi_segments=self.eoi_segments)
                if export_result and export_result.get('format') == 'csv' and export_result.get('reason') == 'openpyxl_not_installed' and hasattr(self, 'progress_update'):
                    self.progress_update.emit("  ⚠ Excel export unavailable (openpyxl missing). CSV fallback used.")
                # Export binned heatmap as JPG instead of PNG
                viz_path = f"{output_prefix}_heatmap.jpg"
                visualize_binned_analysis(binned_data, save_path=viz_path)
        except Exception as e:
            # Non-fatal error: report in progress area
            if hasattr(self, 'progress_update'):
                self.progress_update.emit(f"  ⚠ Binned export failed: {str(e)}")

# =========================================================================== #

class BinnedAnalysisWindow(QDialog):
    '''
        Separate window for binned analysis functionality with dynamic visualization.
        Displays frequency band power, occupancy, and dominant band heatmaps.
    '''
    
    def __init__(self, parent=None, binned_data=None, files=None, active_folder=None, eoi_segments=None, tracking_data=None):
        super().__init__(parent)
        self.binned_data = binned_data
        self.eoi_segments = eoi_segments
        self.tracking_data = tracking_data
        self.files = files or [None, None]
        self.active_folder = active_folder or os.getcwd()
        self.current_chunk = 0
        self.show_percent_power = False
        self.render_timer = QTimer(self)
        self.render_timer.setSingleShot(True)
        self.render_timer.timeout.connect(self.renderChunkViews)
        self.setWindowTitle("Binned Analysis Studio")
        self.setGeometry(100, 100, 1400, 900)
        self.initUI()
    
    def initUI(self):
        '''Initialize the binned analysis window UI'''
        main_layout = QVBoxLayout(self)
        
        # Top section: Title
        self.title_label = QLabel("Binned Frequency Analysis Studio")
        title_label = self.title_label
        title_label.setFont(QFont("Times New Roman", 16, QFont.Bold))
        main_layout.addWidget(title_label)
        
        # Control panel
        control_layout = QHBoxLayout()
        
        # Time chunk slider with step buttons
        slider_label = QLabel("Time Chunk:")
        control_layout.addWidget(slider_label)
        
        prev_btn = QPushButton("←", self)
        prev_btn.setFixedWidth(36)
        prev_btn.clicked.connect(partial(self.stepChunk, -1))
        control_layout.addWidget(prev_btn)
        
        self.chunk_slider = QSlider(Qt.Horizontal)
        self.chunk_slider.setMinimum(0)
        if self.binned_data:
            self.chunk_slider.setMaximum(max(0, self.binned_data['time_chunks'] - 1))
        self.chunk_slider.setSingleStep(1)
        self.chunk_slider.setPageStep(1)
        self.chunk_slider.setValue(0)
        self.chunk_slider.setMaximumWidth(260)
        self.chunk_slider.valueChanged.connect(self.onChunkChanged)
        control_layout.addWidget(self.chunk_slider)
        
        next_btn = QPushButton("→", self)
        next_btn.setFixedWidth(36)
        next_btn.clicked.connect(partial(self.stepChunk, 1))
        control_layout.addWidget(next_btn)
        
        self.chunk_display_label = QLabel("0 / 1")
        self.chunk_display_label.setMinimumWidth(70)
        control_layout.addWidget(self.chunk_display_label)
        
        # Power mode toggle
        control_layout.addSpacing(20)
        self.power_mode_btn = QPushButton("Switch to % Power", self)
        self.power_mode_btn.setCheckable(True)
        self.power_mode_btn.setChecked(False)
        self.power_mode_btn.toggled.connect(self.onPowerModeToggled)
        self.power_mode_btn.setMaximumWidth(150)
        control_layout.addWidget(self.power_mode_btn)
        
        # Buttons
        control_layout.addSpacing(20)
        save_pngs_btn = QPushButton("Export Plots", self)
        save_pngs_btn.clicked.connect(self.exportAllPngs)
        control_layout.addWidget(save_pngs_btn)
        
        export_data_btn = QPushButton("Export Data", self)
        export_data_btn.clicked.connect(self.exportData)
        control_layout.addWidget(export_data_btn)
        
        control_layout.addStretch()
        close_btn = QPushButton("Close", self)
        close_btn.clicked.connect(self.close)
        control_layout.addWidget(close_btn)
        
        main_layout.addLayout(control_layout)
        
        # Display areas with scroll
        display_layout = QHBoxLayout()
        
        # Left column: Mean power
        left_layout = QVBoxLayout()
        left_label = QLabel("Frequency Band Power & Occupancy")
        left_label.setFont(QFont("", 11, QFont.Bold))
        left_layout.addWidget(left_label)
        
        self.power_scroll = QScrollArea(self)
        self.power_scroll.setWidgetResizable(True)
        self.power_label = QLabel()
        self.power_label.setAlignment(Qt.AlignCenter)
        self.power_scroll.setWidget(self.power_label)
        left_layout.addWidget(self.power_scroll)
        display_layout.addLayout(left_layout, 2)
        
        # Right column: Dominant Band & EOI Distribution
        right_layout = QVBoxLayout()
        right_label = QLabel("Dominant Band & EOI Distribution")
        right_label.setFont(QFont("", 11, QFont.Bold))
        right_layout.addWidget(right_label)
        
        self.occ_scroll = QScrollArea(self)
        self.occ_scroll.setWidgetResizable(True)
        self.occ_label = QLabel()
        self.occ_label.setAlignment(Qt.AlignCenter)
        self.occ_scroll.setWidget(self.occ_label)
        right_layout.addWidget(self.occ_scroll)
        display_layout.addLayout(right_layout, 1)
        
        main_layout.addLayout(display_layout, 1)
        
        # Status bar
        self.status_label = QLabel("Ready")
        self.status_label.setStyleSheet("color: gray; font-style: italic;")
        main_layout.addWidget(self.status_label)
        
        self.setLayout(main_layout)
        
        # Initial render
        if self.binned_data:
            self.renderChunkViews()
            
        if self.binned_data and self.binned_data.get('type') == 'polar':
            self.title_label.setText("Polar Binned Analysis (16 Bins)")
        else:
            self.title_label.setText("4×4 Binned Frequency Analysis Studio")
    
    def getOutputFolder(self):
        '''Get or create the binned analysis output folder'''
        out_dir = self.active_folder
        binned_folder = os.path.join(out_dir, "binned_analysis_output")
        if not os.path.exists(binned_folder):
            os.makedirs(binned_folder)
        return binned_folder
    
    def onChunkChanged(self, value):
        '''Update chunk display and re-render on slider change'''
        self.current_chunk = value
        max_chunks = self.binned_data['time_chunks'] if self.binned_data else 1
        # Show chunk index and time range for consistency with main GUI
        if self.binned_data and 'chunk_size' in self.binned_data:
            cs = self.binned_data['chunk_size']
            t_start = value * cs
            # Clamp end to actual duration for the last chunk
            duration = float(self.binned_data.get('duration', (max_chunks * cs)))
            t_end_nominal = (value + 1) * cs
            t_end = min(t_end_nominal, duration)
            self.chunk_display_label.setText(f"Chunk {value+1}/{max_chunks} ({t_start:.0f}-{t_end:.0f}s)")
        else:
            self.chunk_display_label.setText(f"Chunk {value+1}/{max_chunks}")
        # Debounce rendering to keep slider smooth
        self.render_timer.start(60)

    def stepChunk(self, delta):
        '''Increment/decrement the chunk slider and render'''
        if self.binned_data is None:
            return
        max_chunks = self.binned_data['time_chunks'] if self.binned_data else 1
        new_val = int(np.clip(self.chunk_slider.value() + delta, 0, max_chunks - 1))
        self.chunk_slider.setValue(new_val)
    
    def onPowerModeToggled(self, checked):
        '''Toggle between absolute and percent power'''
        self.show_percent_power = checked
        mode = "% Power" if checked else "Absolute Power"
        self.power_mode_btn.setText(f"Switch to {'Absolute' if checked else '%'} Power")
        self.status_label.setText(f"Switched to {mode}")
        self.renderChunkViews()
    
    def renderChunkViews(self):
        '''Render all visualizations for current chunk (in-memory, no temp files)'''
        if self.binned_data is None:
            return
        
        try:
            # Render frequency band power heatmap in memory
            if self.binned_data.get('type') == 'polar':
                fig_power, _ = self._create_polar_power_heatmap(self.current_chunk, self.show_percent_power)
            else:
                fig_power, _ = self._create_power_heatmap(self.current_chunk, self.show_percent_power)
            power_pixmap = self._fig_to_pixmap(fig_power)
            if not power_pixmap.isNull():
                if power_pixmap.width() > 1000:
                    power_pixmap = power_pixmap.scaledToWidth(1000, Qt.SmoothTransformation)
                self.power_label.setPixmap(power_pixmap)
            
            # Render Dominant Band & EOI in memory
            if self.binned_data.get('type') == 'polar':
                fig_dom_eoi, _ = self._create_polar_dominant_eoi_heatmap(self.current_chunk)
            else:
                fig_dom_eoi, _ = self._create_dominant_eoi_heatmap(self.current_chunk)
            occ_pixmap = self._fig_to_pixmap(fig_dom_eoi)
            if not occ_pixmap.isNull():
                if occ_pixmap.width() > 600:
                    occ_pixmap = occ_pixmap.scaledToWidth(600, Qt.SmoothTransformation)
                self.occ_label.setPixmap(occ_pixmap)
            
            self.status_label.setText(f"Chunk {self.current_chunk + 1} rendered")
        except Exception as e:
            self.status_label.setText(f"Error rendering views: {str(e)}")
    
    def _fig_to_pixmap(self, fig):
        '''Convert matplotlib figure to QPixmap without saving to disk'''
        from io import BytesIO
        buf = BytesIO()
        fig.savefig(buf, format='png', dpi=100, bbox_inches='tight')
        buf.seek(0)
        pixmap = QPixmap()
        pixmap.loadFromData(buf.getvalue(), 'PNG')
        plt.close(fig)
        return pixmap
    
    def _get_bounds(self):
        '''Calculate global bounds from tracking data'''
        if self.tracking_data:
            all_x = np.concatenate(self.tracking_data[0])
            all_y = np.concatenate(self.tracking_data[1])
            if len(all_x) > 0:
                return np.min(all_x), np.max(all_x), np.min(all_y), np.max(all_y)
        return 0, 1, 0, 1

    def _create_power_heatmap(self, chunk_idx, show_percent=False):
        '''Create power heatmap figure (does not save)'''
        n_chunks = self.binned_data['time_chunks']
        if chunk_idx < 0 or chunk_idx >= n_chunks:
            chunk_idx = max(0, min(chunk_idx, n_chunks - 1))
        
        fig, axes = plt.subplots(2, 4, figsize=(16, 8))
        power_type = "Percent Power" if show_percent else "Power"
        fig.suptitle(f'4x4 Spatial Bins - Chunk {chunk_idx + 1} (Frequency Band {power_type} & Occupancy)', 
                     fontsize=14, fontweight='bold')
        
        bands = self.binned_data['bands']
        
        # Get min/max across all chunks for consistent color scale
        vmin_all = {}
        vmax_all = {}
        for band in bands:
            timeseries_data = self.binned_data['bin_power_timeseries'][band]
            
            if show_percent:
                percent_data = np.zeros_like(timeseries_data)
                for t in range(timeseries_data.shape[2]):
                    for x in range(4):
                        for y in range(4):
                            total_power = sum(self.binned_data['bin_power_timeseries'][b][x, y, t] 
                                            for b in bands)
                            if total_power > 0:
                                percent_data[x, y, t] = (timeseries_data[x, y, t] / total_power) * 100
                vmin_all[band] = np.nanmin(percent_data)
                vmax_all[band] = np.nanmax(percent_data)
            else:
                vmin_all[band] = np.nanmin(timeseries_data)
                vmax_all[band] = np.nanmax(timeseries_data)
        
        # First row: First 4 bands
        for idx, band in enumerate(bands[:4]):
            chunk_power = self.binned_data['bin_power_timeseries'][band][:, :, chunk_idx]
            
            if show_percent:
                total_power = sum(self.binned_data['bin_power_timeseries'][b][:, :, chunk_idx] 
                                for b in bands)
                chunk_power = np.divide(chunk_power, total_power, where=total_power>0, 
                                       out=np.zeros_like(chunk_power)) * 100
            
            im = axes[0, idx].imshow(chunk_power, cmap=GUI_CMAP_POWER, aspect='equal', interpolation='nearest',
                                     vmin=vmin_all[band], vmax=vmax_all[band])
            axes[0, idx].set_title(f'{band}')
            axes[0, idx].set_xticks([0, 1, 2, 3])
            axes[0, idx].set_yticks([0, 1, 2, 3])
            axes[0, idx].grid(True, alpha=0.3)
            cbar = plt.colorbar(im, ax=axes[0, idx])
            if not show_percent:
                cbar.ax.yaxis.set_major_formatter(FuncFormatter(lambda x, pos: f'{x*1e-3:g}K' if x >= 1000 else f'{x:g}'))
            cbar.set_label('%' if show_percent else 'Power', fontsize=9)
        
        # Second row: Remaining bands
        remaining_bands = bands[4:]
        for idx, band in enumerate(remaining_bands):
            chunk_power = self.binned_data['bin_power_timeseries'][band][:, :, chunk_idx]
            
            if show_percent:
                total_power = sum(self.binned_data['bin_power_timeseries'][b][:, :, chunk_idx] 
                                for b in bands)
                chunk_power = np.divide(chunk_power, total_power, where=total_power>0, 
                                       out=np.zeros_like(chunk_power)) * 100
            im = axes[1, idx].imshow(chunk_power, cmap=GUI_CMAP_POWER, aspect='equal', interpolation='nearest',
                                     vmin=vmin_all[band], vmax=vmax_all[band])
            axes[1, idx].set_title(f'{band}')
            axes[1, idx].set_xticks([0, 1, 2, 3])
            axes[1, idx].set_yticks([0, 1, 2, 3])
            axes[1, idx].grid(True, alpha=0.3)
            cbar = plt.colorbar(im, ax=axes[1, idx])
            if not show_percent:
                cbar.ax.yaxis.set_major_formatter(FuncFormatter(lambda x, pos: f'{x*1e-3:g}K' if x >= 1000 else f'{x:g}'))
            cbar.set_label('%' if show_percent else 'Power', fontsize=9)
        
        # 8th slot: Occupancy (Dynamic per chunk)
        ax_occ = axes[1, 3]
        occ_data = None
        
        # Try to get pre-calculated timeseries (Polar usually has this)
        if 'bin_occupancy_timeseries' in self.binned_data:
             occ_data = self.binned_data['bin_occupancy_timeseries'][:, :, chunk_idx]
        # Fallback: Calculate from tracking data for 4x4 if missing
        elif self.tracking_data:
             min_x, max_x, min_y, max_y = self._get_bounds()
             if chunk_idx < len(self.tracking_data[0]):
                 cx = self.tracking_data[0][chunk_idx]
                 cy = self.tracking_data[1][chunk_idx]
                 x_edges = np.linspace(min_x, max_x, 5)
                 y_edges = np.linspace(min_y, max_y, 5)
                 H, _, _ = np.histogram2d(cx, cy, bins=[x_edges, y_edges])
                 occ_data = np.flipud(H.T)
        
        if occ_data is not None:
             total = np.sum(occ_data)
             if total > 0: occ_data = (occ_data / total) * 100.0
             im = ax_occ.imshow(occ_data, cmap=GUI_CMAP_OCCUPANCY, aspect='equal', interpolation='nearest', vmin=0, vmax=100)
             ax_occ.set_title('Occupancy (%)')
             ax_occ.set_xticks([0, 1, 2, 3])
             ax_occ.set_yticks([0, 1, 2, 3])
             ax_occ.grid(True, alpha=0.3)
             cbar = plt.colorbar(im, ax=ax_occ)
             cbar.set_label('%', fontsize=9)
        else:
             ax_occ.text(0.5, 0.5, "No Data", ha='center', va='center')
             ax_occ.axis('off')
        
        plt.tight_layout()
        return fig, axes
    
    def _create_polar_power_heatmap(self, chunk_idx, show_percent=False):
        '''Create polar power heatmap figure'''
        n_chunks = self.binned_data['time_chunks']
        if chunk_idx < 0 or chunk_idx >= n_chunks:
            chunk_idx = max(0, min(chunk_idx, n_chunks - 1))
            
        fig, axes = plt.subplots(2, 4, figsize=(16, 8), subplot_kw={'projection': 'polar'})
        power_type = "Percent Power" if show_percent else "Power"
        fig.suptitle(f'Polar Bins - Chunk {chunk_idx + 1} (Frequency Band {power_type} & Occupancy)', 
                     fontsize=14, fontweight='bold')
        
        bands = self.binned_data['bands']
        
        # Setup grid
        theta = np.linspace(-np.pi, np.pi, 9)
        r = [0, 1.0/np.sqrt(2), 1]
        T, R = np.meshgrid(theta, r)
        
        # Flatten axes
        ax_flat = axes.flatten()
        
        for idx, band in enumerate(bands):
            if idx >= len(ax_flat): break
            ax = ax_flat[idx]
            
            # Data shape (2, 8) -> (Rings, Sectors)
            data = self.binned_data['bin_power_timeseries'][band][:, :, chunk_idx]
            
            if show_percent:
                total_power = sum(self.binned_data['bin_power_timeseries'][b][:, :, chunk_idx] 
                                for b in bands)
                data = np.divide(data, total_power, where=total_power>0, 
                                       out=np.zeros_like(data)) * 100
            
            # pcolormesh expects data to match grid cells
            # T, R shape is (3, 9). Data shape is (2, 8). Perfect.
            if show_percent:
                im = ax.pcolormesh(T, R, data, cmap=GUI_CMAP_PERCENT, shading='flat', vmin=0, vmax=100)
            else:
                im = ax.pcolormesh(T, R, data, cmap=GUI_CMAP_POWER, shading='flat')
            
            ax.set_title(f'{band}')
            ax.set_yticklabels([])
            ax.set_xticklabels([])
            ax.grid(True, alpha=0.3)
            
            # Add colorbar
            cbar = plt.colorbar(im, ax=ax, pad=0.05, shrink=0.8)
            cbar.set_label('%' if show_percent else 'Power', fontsize=8)

        # 8th slot: Occupancy
        if len(bands) < 8:
            ax_occ = ax_flat[7]
            occ_data = None
            if 'bin_occupancy_timeseries' in self.binned_data:
                 occ_data = self.binned_data['bin_occupancy_timeseries'][:, :, chunk_idx]
            
            if occ_data is not None:
                 total = np.sum(occ_data)
                 if total > 0: occ_data = (occ_data / total) * 100.0
                 im = ax_occ.pcolormesh(T, R, occ_data, cmap=GUI_CMAP_OCCUPANCY, shading='flat', vmin=0, vmax=100)
                 ax_occ.set_title('Occupancy (%)')
                 ax_occ.set_yticklabels([])
                 ax_occ.set_xticklabels([])
                 ax_occ.grid(True, alpha=0.3)
                 cbar = plt.colorbar(im, ax=ax_occ, pad=0.05, shrink=0.8)
                 cbar.set_label('%', fontsize=8)
            else:
                 ax_occ.axis('off')
            
        plt.tight_layout()
        return fig, axes

    def _create_dominant_eoi_heatmap(self, chunk_idx):
        '''Create Dominant Band and EOI Distribution heatmap'''
        n_chunks = self.binned_data['time_chunks']
        if chunk_idx < 0 or chunk_idx >= n_chunks:
            chunk_idx = max(0, min(chunk_idx, n_chunks - 1))
        
        fig, axes = plt.subplots(1, 2, figsize=(12, 5))
        
        # Left panel: Dominant band for specific chunk
        dominant_chunk = self.binned_data['bin_dominant_band'][chunk_idx]
        bands = self.binned_data['bands']
        band_map = {band: idx for idx, band in enumerate(bands)}
        numeric_dominant = np.zeros((4, 4))
        for x in range(4):
            for y in range(4):
                band = dominant_chunk[x, y]
                numeric_dominant[x, y] = band_map.get(band, 0)
        
        im1 = axes[0].imshow(numeric_dominant, cmap=GUI_CMAP_DOMINANT, aspect='equal', interpolation='nearest', vmin=0, vmax=len(bands)-1)
        axes[0].set_title(f'Dominant Band - Chunk {chunk_idx + 1}')
        axes[0].set_xticks([0, 1, 2, 3])
        axes[0].set_yticks([0, 1, 2, 3])
        axes[0].grid(True, alpha=0.3)
        
        # Add colorbar with band labels
        cbar1 = plt.colorbar(im1, ax=axes[0], ticks=range(len(bands)))
        cbar1.set_ticklabels(bands, fontsize=8)
        cbar1.set_label('Band', fontsize=9)
        
        # Right panel: EOI Distribution
        if self.eoi_segments:
            min_x, max_x, min_y, max_y = self._get_bounds()
            eoi_x, eoi_y = [], []
            if chunk_idx in self.eoi_segments:
                for seg in self.eoi_segments[chunk_idx]:
                    eoi_x.extend(seg[0])
                    eoi_y.extend(seg[1])
            
            x_edges = np.linspace(min_x, max_x, 5)
            y_edges = np.linspace(min_y, max_y, 5)
            if eoi_x:
                H, _, _ = np.histogram2d(eoi_x, eoi_y, bins=[x_edges, y_edges])
                H = np.flipud(H.T)
            else:
                H = np.zeros((4, 4))
            
            im2 = axes[1].imshow(H, cmap=GUI_CMAP_OCCUPANCY, aspect='equal', interpolation='nearest')
            axes[1].set_title(f'EOI Distribution - Chunk {chunk_idx + 1}')
            axes[1].set_xticks([0, 1, 2, 3])
            axes[1].set_yticks([0, 1, 2, 3])
            axes[1].grid(True, alpha=0.3)
            cbar2 = plt.colorbar(im2, ax=axes[1])
            cbar2.set_label('EOI Count', fontsize=9)
        else:
            axes[1].axis('off')
        
        plt.tight_layout()
        return fig, axes
    
    def _create_polar_dominant_eoi_heatmap(self, chunk_idx):
        '''Create polar Dominant Band and EOI Distribution heatmap'''
        n_chunks = self.binned_data['time_chunks']
        if chunk_idx < 0 or chunk_idx >= n_chunks:
            chunk_idx = max(0, min(chunk_idx, n_chunks - 1))
        
        fig, axes = plt.subplots(1, 2, figsize=(12, 5), subplot_kw={'projection': 'polar'})
        
        # Setup grid
        theta = np.linspace(-np.pi, np.pi, 9)
        r = [0, 1.0/np.sqrt(2), 1]
        T, R = np.meshgrid(theta, r)
        
        # Left: Dominant Band
        dominant_chunk = self.binned_data['bin_dominant_band'][chunk_idx]
        bands = self.binned_data['bands']
        band_map = {band: idx for idx, band in enumerate(bands)}
        
        numeric_dominant = np.zeros((2, 8))
        for r_idx in range(2):
            for s_idx in range(8):
                band = dominant_chunk[r_idx, s_idx]
                numeric_dominant[r_idx, s_idx] = band_map.get(band, 0)
        
        im1 = axes[0].pcolormesh(T, R, numeric_dominant, cmap=GUI_CMAP_DOMINANT, shading='flat',
                                 vmin=0, vmax=len(bands)-1)
        axes[0].set_title(f'Dominant Band - Chunk {chunk_idx + 1}')
        axes[0].set_yticklabels([])
        
        cbar1 = plt.colorbar(im1, ax=axes[0], ticks=range(len(bands)), pad=0.1)
        cbar1.set_ticklabels(bands, fontsize=8)
        
        # Right: EOI Distribution
        if self.eoi_segments:
            min_x, max_x, min_y, max_y = self._get_bounds()
            width = max_x - min_x if max_x > min_x else 1
            height = max_y - min_y if max_y > min_y else 1
            
            H = np.zeros((2, 8))
            if chunk_idx in self.eoi_segments:
                for seg in self.eoi_segments[chunk_idx]:
                    nx = 2 * (np.array(seg[0]) - min_x) / width - 1
                    ny = 2 * (np.array(seg[1]) - min_y) / height - 1
                    r_pt = np.sqrt(nx**2 + ny**2)
                    theta_pt = np.arctan2(ny, nx)
                    
                    equal_area_radius = 1.0 / np.sqrt(2.0)
                    r_bins = [0, equal_area_radius, np.inf]
                    r_indices = np.clip(np.digitize(r_pt, r_bins) - 1, 0, 1)
                    
                    theta_edges = np.linspace(-np.pi, np.pi, 9)
                    th_indices = np.clip(np.digitize(theta_pt, theta_edges) - 1, 0, 7)
                    
                    for ri, ti in zip(r_indices, th_indices):
                        H[ri, ti] += 1
            
            im2 = axes[1].pcolormesh(T, R, H, cmap=GUI_CMAP_OCCUPANCY, shading='flat')
            axes[1].set_title(f'EOI Distribution - Chunk {chunk_idx + 1}')
            axes[1].set_yticklabels([])
            cbar2 = plt.colorbar(im2, ax=axes[1], pad=0.1)
            cbar2.set_label('EOI Count', fontsize=9)
        else:
            axes[1].axis('off')
        
        plt.tight_layout()
        return fig, axes

    def updateBinnedData(self, binned_data, files=None, active_folder=None, eoi_segments=None, tracking_data=None):
        '''Update binned data when new session is loaded'''
        self.binned_data = binned_data
        self.eoi_segments = eoi_segments
        self.tracking_data = tracking_data
        if files:
            self.files = files
        if active_folder:
            self.active_folder = active_folder
        
        # Update slider range
        if binned_data:
            max_chunks = binned_data['time_chunks']
            self.chunk_slider.setMaximum(max(0, max_chunks - 1))
            self.chunk_slider.setValue(0)
            self.current_chunk = 0
            if 'chunk_size' in binned_data:
                cs = binned_data['chunk_size']
                duration = float(binned_data.get('duration', (max_chunks * cs)))
                t_end = min(cs, duration)
                self.chunk_display_label.setText(f"Chunk 1/{max_chunks} (0-{t_end:.0f}s)")
            else:
                self.chunk_display_label.setText(f"Chunk 1/{max_chunks}")
            self.renderChunkViews()
            
            if binned_data.get('type') == 'polar':
                self.title_label.setText("Polar Binned Analysis (16 Bins)")
            else:
                self.title_label.setText("4×4 Binned Frequency Analysis Studio")
        
        self.status_label.setText("Data updated. Ready for analysis.")
    
    def exportAllPngs(self):
        '''Export all JPG visualizations (Power+Occupancy, Dominant+EOI per chunk)'''
        if self.binned_data is None:
            QMessageBox.information(self, 'Export PNGs', 'No binned data available.')
            return
        
        try:
            self.status_label.setText("Exporting all JPGs...")
            QApplication.processEvents()
            
            output_folder = self.getOutputFolder()
            base_name = os.path.splitext(os.path.basename(self.files[1] or 'output'))[0]
            
            n_chunks = self.binned_data['time_chunks']
            export_count = 0
            
            # Export mean power + occupancy for all chunks
            for chunk_idx in range(n_chunks):
                if self.binned_data.get('type') == 'polar':
                    fig, _ = self._create_polar_power_heatmap(chunk_idx, show_percent=False)
                    jpg_path = os.path.join(output_folder, f"{base_name}_chunk{chunk_idx+1:02d}_polar_power.jpg")
                else:
                    fig, _ = self._create_power_heatmap(chunk_idx, show_percent=False)
                    jpg_path = os.path.join(output_folder, f"{base_name}_chunk{chunk_idx+1:02d}_mean_power.jpg")
                fig.savefig(jpg_path, format='jpg', pil_kwargs={'quality': 85}, bbox_inches='tight')
                plt.close(fig)
                export_count += 1
            
            # Export percent power + occupancy for all chunks
            for chunk_idx in range(n_chunks):
                if self.binned_data.get('type') == 'polar':
                    fig, _ = self._create_polar_power_heatmap(chunk_idx, show_percent=True)
                    jpg_path = os.path.join(output_folder, f"{base_name}_chunk{chunk_idx+1:02d}_polar_percent.jpg")
                else:
                    fig, _ = self._create_power_heatmap(chunk_idx, show_percent=True)
                    jpg_path = os.path.join(output_folder, f"{base_name}_chunk{chunk_idx+1:02d}_percent_power.jpg")
                fig.savefig(jpg_path, format='jpg', pil_kwargs={'quality': 85}, bbox_inches='tight')
                plt.close(fig)
                export_count += 1
            
            # Export dominant band + EOI per chunk
            for chunk_idx in range(n_chunks):
                if self.binned_data.get('type') == 'polar':
                    fig, _ = self._create_polar_dominant_eoi_heatmap(chunk_idx)
                    jpg_path = os.path.join(output_folder, f"{base_name}_chunk{chunk_idx+1:02d}_polar_dominant_eoi.jpg")
                else:
                    fig, _ = self._create_dominant_eoi_heatmap(chunk_idx)
                    jpg_path = os.path.join(output_folder, f"{base_name}_chunk{chunk_idx+1:02d}_dominant_eoi.jpg")
                fig.savefig(jpg_path, format='jpg', pil_kwargs={'quality': 85}, bbox_inches='tight')
                plt.close(fig)
                export_count += 1
            
            self.status_label.setText(f"✓ Exported {export_count} JPGs")
            QMessageBox.information(
                self,
                'Export Complete',
                f'Exported {export_count} JPG files to:\n{output_folder}\n\n'
                f'  • {n_chunks} mean power & occupancy heatmaps\n'
                f'  • {n_chunks} percent power & occupancy heatmaps\n'
                f'  • {n_chunks} dominant band & EOI heatmaps\n\n'
                f'(JPG format for faster export & smaller file size)'
            )
        except Exception as e:
            self.status_label.setText(f"Export failed: {str(e)}")
            QMessageBox.warning(self, 'Export Error', f'Failed to export JPGs: {str(e)}')
    
    def exportData(self):
        '''Export binned analysis data to Excel files (or CSV fallback)'''
        if self.binned_data is None:
            QMessageBox.information(self, 'Export Data', 'No binned data available. Please run a session first.')
            return
        
        try:
            self.status_label.setText("Exporting data...")
            QApplication.processEvents()
            
            output_folder = self.getOutputFolder()
            base_name = os.path.splitext(os.path.basename(self.files[1] or 'output'))[0]
            
            if self.binned_data.get('type') == 'polar':
                # Polar data export
                output_prefix = os.path.join(output_folder, f"{base_name}_binned")
                
                try:
                    import openpyxl
                    use_excel = True
                except ImportError:
                    use_excel = False
                
                # Prepare data
                bands = self.binned_data['bands']
                n_chunks = self.binned_data['time_chunks']
                
                power_per_chunk = {}
                percent_per_chunk = {}
                
                # Helper to reshape (2, 8, n_chunks) -> (n_chunks, 16) and add Chunk column
                def prepare_matrix(data_3d):
                    # reshape to (16, n_chunks) -> transpose to (n_chunks, 16)
                    flat = data_3d.reshape(16, -1).T
                    # Add chunk column (1-based)
                    chunks_col = np.arange(1, n_chunks + 1).reshape(-1, 1)
                    return np.hstack([chunks_col, flat])
                
                # Calculate total power per bin per chunk for percentage
                all_bands_power = np.stack([self.binned_data['bin_power_timeseries'][b] for b in bands])
                total_power = np.sum(all_bands_power, axis=0) # (2, 8, n_chunks)
                
                for band in bands:
                    # Power per chunk
                    data = self.binned_data['bin_power_timeseries'][band]
                    power_per_chunk[band] = prepare_matrix(data)
                    
                    # Percent per chunk
                    with np.errstate(divide='ignore', invalid='ignore'):
                        pct_data = np.where(total_power > 0, (data / total_power) * 100.0, 0.0)
                    percent_per_chunk[band] = prepare_matrix(pct_data)
                
                # Dominant band counts
                dominant_counts = {band: np.zeros((2, 8)) for band in bands}
                for chunk_data in self.binned_data['bin_dominant_band']:
                    for r_idx in range(chunk_data.shape[0]):
                        for s_idx in range(chunk_data.shape[1]):
                            band = chunk_data[r_idx, s_idx]
                            if band in dominant_counts:
                                dominant_counts[band][r_idx, s_idx] += 1
                
                # Occupancy per chunk (Percent)
                occupancy_per_chunk = None
                if 'bin_occupancy_timeseries' in self.binned_data:
                    occ_ts = self.binned_data['bin_occupancy_timeseries']
                    # Calculate percent occupancy per chunk
                    occ_sums = np.sum(occ_ts, axis=(0, 1))
                    with np.errstate(divide='ignore', invalid='ignore'):
                        # Broadcasting: (2,8,n) / (n) -> need (1,1,n) for broadcasting or transpose
                        occ_pct = np.where(occ_sums > 0, (occ_ts / occ_sums[None, None, :]) * 100.0, 0.0)
                    occupancy_per_chunk = prepare_matrix(occ_pct)

                # EOI Distribution (Per Chunk)
                eoi_counts_per_chunk = np.zeros((2, 8, n_chunks))
                eoi_export_matrix = None
                
                if self.eoi_segments and self.tracking_data:
                    eoi_counts_per_chunk = np.zeros((2, 8, n_chunks))
                    # Calculate bounds (same as in _create_polar_eoi_heatmap)
                    all_x = np.concatenate(self.tracking_data[0])
                    all_y = np.concatenate(self.tracking_data[1])
                    min_x, max_x = np.min(all_x), np.max(all_x)
                    min_y, max_y = np.min(all_y), np.max(all_y)
                    width = max_x - min_x
                    height = max_y - min_y
                    if width == 0: width = 1
                    if height == 0: height = 1
                    
                    # Iterate all segments
                    for chunk_idx, segments in self.eoi_segments.items():
                        for seg in segments:
                            nx = 2 * (np.array(seg[0]) - min_x) / width - 1
                            ny = 2 * (np.array(seg[1]) - min_y) / height - 1
                            r = np.sqrt(nx**2 + ny**2)
                            theta = np.arctan2(ny, nx)
                            
                            equal_area_radius = 1.0 / np.sqrt(2.0)
                            r_bins = [0, equal_area_radius, np.inf]
                            r_indices = np.clip(np.digitize(r, r_bins) - 1, 0, 1)
                            
                            theta_edges = np.linspace(-np.pi, np.pi, 9)
                            th_indices = np.clip(np.digitize(theta, theta_edges) - 1, 0, 7)
                            
                            for ri, ti in zip(r_indices, th_indices):
                                eoi_counts_per_chunk[ri, ti, chunk_idx] += 1
                    eoi_export_matrix = prepare_matrix(eoi_counts_per_chunk)
                
                # 5. Dominant Band (per chunk)
                dom_data = np.array(self.binned_data['bin_dominant_band']) # (n_chunks, 2, 8)
                dom_data_T = dom_data.transpose(1, 2, 0) # (2, 8, n_chunks)
                dom_export_matrix = prepare_matrix(dom_data_T)
                
                exported_files = []
                
                if use_excel:
                    bin_headers = ["Chunk"] + [f"Inner_S{i+1}" for i in range(8)] + [f"Outer_S{i+1}" for i in range(8)]

                    # Helper to save multi-sheet workbook
                    def save_multi_sheet(data_dict, filename):
                        wb = openpyxl.Workbook()
                        wb.remove(wb.active)
                        for sheet_name, matrix in data_dict.items():
                            ws = wb.create_sheet(title=sheet_name)
                            ws.append(bin_headers)
                            for row in matrix:
                                ws.append([float(val) for val in row])
                        wb.save(filename)
                        exported_files.append(filename)

                    # Helper to save single-sheet workbook
                    def save_single_sheet(matrix, filename, sheet_title, is_string=False):
                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.title = sheet_title
                        ws.append(bin_headers)
                        for row in matrix:
                            if is_string:
                                ws.append([row[0]] + list(row[1:]))
                            else:
                                ws.append([float(val) for val in row])
                        wb.save(filename)
                        exported_files.append(filename)
                    
                    # 1. Mean Power
                    save_multi_sheet(power_per_chunk, f"{output_prefix}_mean_power_per_chunk.xlsx")
                    # 2. Percent Power
                    save_multi_sheet(percent_per_chunk, f"{output_prefix}_percent_power_per_chunk.xlsx")
                    
                    # 3. EOIs
                    if eoi_export_matrix is not None:
                        save_single_sheet(eoi_export_matrix, f"{output_prefix}_eois_per_chunk.xlsx", "EOIs")
                    
                    # 4. Percent Occupancy
                    if occupancy_per_chunk is not None:
                        save_single_sheet(occupancy_per_chunk, f"{output_prefix}_percent_occupancy_per_chunk.xlsx", "Percent Occupancy")
                    
                    # 5. Dominant Band
                    save_single_sheet(dom_export_matrix, f"{output_prefix}_dominant_band_per_chunk.xlsx", "Dominant Band", is_string=True)
                    
                    format_str = "EXCEL"
                else:
                    # CSV Fallback
                    for band in bands:
                        np.savetxt(f"{output_prefix}_mean_power_{band}.csv", power_per_chunk[band], delimiter=',')
                        np.savetxt(f"{output_prefix}_percent_power_{band}.csv", percent_per_chunk[band], delimiter=',')
                    
                    if eoi_export_matrix is not None:
                        np.savetxt(f"{output_prefix}_eois.csv", eoi_export_matrix, delimiter=',')
                    
                    if occupancy_per_chunk is not None:
                        np.savetxt(f"{output_prefix}_percent_occupancy.csv", occupancy_per_chunk, delimiter=',')
                    
                    np.savetxt(f"{output_prefix}_dominant_band.csv", dom_export_matrix, delimiter=',', fmt='%s')
                    
                    format_str = "CSV"
                    
                    # Warn about openpyxl
                    warn = QMessageBox(self)
                    warn.setWindowTitle('Excel Export Unavailable')
                    warn.setText('openpyxl not installed — exported CSV instead.\n\nInstall openpyxl to enable Excel export with multiple sheets.')
                    install_btn = warn.addButton('Install openpyxl', QMessageBox.AcceptRole)
                    warn.addButton('Skip', QMessageBox.RejectRole)
                    warn.exec_()
                    if warn.clickedButton() == install_btn:
                        try:
                            self.status_label.setText('Installing openpyxl...')
                            QApplication.processEvents()
                            proc = subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl'], capture_output=True, text=True)
                            if proc.returncode == 0:
                                QMessageBox.information(self, 'Installation Complete', 'openpyxl installed successfully. Please try exporting again.')
                                return
                            else:
                                QMessageBox.warning(self, 'Installation Failed', f'Could not install openpyxl.\n\n{proc.stderr[:500]}')
                        except Exception as ie:
                            QMessageBox.warning(self, 'Installation Error', f'Error installing openpyxl: {str(ie)}')

                self.status_label.setText(f"✓ Data exported ({format_str})")
                file_list = "\n".join([f"  • {os.path.basename(p)}" for p in exported_files])
                QMessageBox.information(self, 'Export Complete',
                    f'Binned analysis data exported to:\n{output_folder}\n\n'
                    f'Format: {format_str}\n\nFiles:\n{file_list}')
                return
            else:
                pos_x_chunks = self.tracking_data[0] if self.tracking_data else None
                pos_y_chunks = self.tracking_data[1] if self.tracking_data else None
                result = export_binned_analysis_to_csv(self.binned_data, os.path.join(output_folder, f"{base_name}_binned"), pos_x_chunks=pos_x_chunks, pos_y_chunks=pos_y_chunks, eoi_segments=self.eoi_segments)
            self.status_label.setText(f"✓ Data exported ({result['format']})")
            file_list = "\n".join([f"  • {os.path.basename(p)}" for p in result['files']])
            if result.get('format') == 'csv' and result.get('reason') == 'openpyxl_not_installed':
                warn = QMessageBox(self)
                warn.setWindowTitle('Excel Export Unavailable')
                warn.setText('openpyxl not installed — exported CSV instead.\n\nInstall openpyxl to enable Excel export with multiple sheets.')
                install_btn = warn.addButton('Install openpyxl', QMessageBox.AcceptRole)
                warn.addButton('Skip', QMessageBox.RejectRole)
                warn.exec_()
                if warn.clickedButton() == install_btn:
                    try:
                        self.status_label.setText('Installing openpyxl...')
                        QApplication.processEvents()
                        proc = subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl'], capture_output=True, text=True)
                        if proc.returncode == 0:
                            QMessageBox.information(self, 'Installation Complete', 'openpyxl installed successfully. Re-exporting to Excel...')
                            # Re-run export to produce Excel files
                            pos_x_chunks = self.tracking_data[0] if self.tracking_data else None
                            pos_y_chunks = self.tracking_data[1] if self.tracking_data else None
                            result = export_binned_analysis_to_csv(self.binned_data, os.path.join(output_folder, f"{base_name}_binned"), pos_x_chunks=pos_x_chunks, pos_y_chunks=pos_y_chunks, eoi_segments=self.eoi_segments)
                            self.status_label.setText(f"✓ Data exported ({result['format']})")
                            file_list = "\n".join([f"  • {os.path.basename(p)}" for p in result['files']])
                        else:
                            QMessageBox.warning(self, 'Installation Failed', f'Could not install openpyxl.\n\n{proc.stderr[:500]}')
                    except Exception as ie:
                        QMessageBox.warning(self, 'Installation Error', f'Error installing openpyxl: {str(ie)}')
            QMessageBox.information(self, 'Export Complete',
                f'Binned analysis data exported to:\n{output_folder}\n\n'
                f'Format: {result["format"].upper()}\n\nFiles:\n{file_list}')
        except Exception as e:
            self.status_label.setText(f"Export failed: {str(e)}")
            QMessageBox.warning(self, 'Export Error', f'Failed to export data: {str(e)}')

# =========================================================================== #

class MplCanvas(FigureCanvasQTAgg):

    '''
        Canvas class to generate matplotlib plot widgets in the GUI.
        This class takes care of plotting any image data.
    '''

    def __init__(self, parent=None, width=5, height=4, dpi=100):
        fig = Figure(figsize=(width, height), dpi=dpi)
        self.axes = fig.add_subplot(111)
        super(MplCanvas, self).__init__(fig)

# =========================================================================== #
        
class frequencyPlotWindow(QWidget):
    
    '''
        Class which handles frequency map plotting on UI
    '''

    def __init__(self, eeg_file=None, ppm=None, eoi_file=None):
        
        QWidget.__init__(self)
        self.eeg_file_arg = eeg_file  # Store the EEG file passed from command line
        self.ppm_arg = ppm  # Store the PPM value passed from command line
        self.eoi_file_arg = eoi_file # Store the EOI file passed from command line
        self.eoi_segments = {} # Holds processed EOI segments for plotting
        
        # Setting main window geometry
        self.center()
        self.mainUI()
        
        # Open at 70% of screen resolution instead of fullscreen
        screen = QApplication.primaryScreen().geometry()
        width = int(screen.width() * 0.7)
        height = int(screen.height() * 0.7)
        x = (screen.width() - width) // 2
        y = (screen.height() - height) // 2
        self.setGeometry(x, y, width, height)
        
        self.setWindowFlag(Qt.WindowCloseButtonHint, False) # Grey out 'X' button to prevent improper PyQt termination
        
    # ------------------------------------------- # 
    
    def center(self):

        '''
            Centers the GUI window on screen upon launch
        '''

        # Geometry of the main window
        qr = self.frameGeometry()

        # Center point of screen
        cp = QDesktopWidget().availableGeometry().center()

        # Move rectangle's center point to screen's center point
        qr.moveCenter(cp)

        # Top left of rectangle becomes top left of window centering it
        self.move(qr.topLeft())
        
    # ------------------------------------------- #  
    
    def mainUI(self):
        
        # Initialize layout and title
        self.layout = QGridLayout()
        self.setLayout(self.layout)
        self.setWindowTitle("Spatial Map")
        
        # Data initialization
        self.plot_flag = False          # Flag to signal if we are plotting graphs or maps. True is graph, false is maps
        self.plot_data = None           # Holds aray of power spectrum densities and frequency axis from welch method
        self.frequencyBand = 'Delta'    # Current band of frequency (choose between Delta, Theta, Beta, Low Gamma, High Gamma)
        self.files = [None, None]       # Holds .pos and .eeg/.egf file
        self.position_data = [None, None, None] # Hods pos_x, pos_y and pos_t
        self.active_folder = ''                 # Keeps track of last accessed directory
        self.ppm = 511                          # Pixel per meter value
        self.chunk_size = 60                    # Size of each signal chunk in seconds (user defined)
        self.window_type = 'hann'               # Window type for welch 
        self.speed_lowerbound = None            # Lower limit for speed filter
        self.speed_upperbound = None            # Upper limit for speed filter
        self.images = None                      # Holds all frequency map images
        self.freq_dict = None                   # Holds frequency map bands and associated Hz ranges
        self.pos_t = None                       # Time tracking array
        self.scaling_factor_crossband = None    # Will hold scaling factor to normalize maps across all bands
        self.chunk_powers_data = None           # Array of total powers per chunk per band
        self.chunk_index = None                 # Keeps track of what signal chunk we are on
        self.tracking_data = None               # Keeps track of animal's position data (x and y coordinates) for plotting
         
        # Widget initialization
        windowTypeBox = QComboBox()
        
        timeSlider_Label = QLabel()
        ppm_Label = QLabel()
        chunkSize_Label = QLabel()
        speed_Label = QLabel()
        window_Label = QLabel()
        session_Label = QLabel()
        self.timeInterval_Label = QLabel()
        self.session_Text = QLabel()
        self.progressBar_Label = QLabel()
        self.power_Label = QTextEdit()
        self.power_Label.setReadOnly(True)
        self.power_Label.setWordWrapMode(True)
        self.power_Label.setMaximumHeight(150)
        self.frequencyViewer_Label = QLabel()
        self.graph_Label = QLabel()
        self.tracking_Label = QLabel()
        self.binned_analysis_window = None  # Reference to separate binned analysis window
        
        self.ppmTextBox = QLineEdit(self)
        chunkSizeTextBox = QLineEdit(self)
        speedTextBox = QLineEdit()
        quit_button = QPushButton('Quit', self)
        self.browse_button = QPushButton('Browse file', self)
        self.browse_button_default_style = self.browse_button.styleSheet()
        browse_folder_button = QPushButton('Browse folder', self)
        self.binned_analysis_btn = QPushButton('Binned Analysis Studio', self)
        self.graph_mode_button = QPushButton('Graph mode', self)
        self.render_button = QPushButton('Re-Render', self)
        save_button = QPushButton('Save data', self)
        self.slider = QSlider(Qt.Horizontal)
        self.full_duration_btn = QPushButton('Full Duration', self)
        self.full_duration_btn.setCheckable(True)
        self.full_duration_btn.clicked.connect(self.toggleFullDuration)
        self.full_duration_btn.setFixedWidth(100)
        self.load_eoi_btn = QPushButton('Load EOIs', self)
        self.bar = QProgressBar(self)
        
        # Create canvases for embedded plotting
        self.graph_canvas = MplCanvas(self, width=5, height=5, dpi=100)     # For fft plotting
        self.tracking_canvas = MplCanvas(self, width=6, height=6, dpi=100)  # For position tracking plotting

        self.scene = QGraphicsScene(self)
        self.view = QGraphicsView(self.scene)
        self.imageMapper = QGraphicsPixmapItem()
        self.scene.addItem(self.imageMapper)
        self.view.centerOn(self.imageMapper)
        self.view.scale(3,3)
        # Set maximum size for frequency map view to match tracking canvas size
        self.view.setMaximumSize(600, 600)  # Similar to tracking canvas (6 inches * 100 dpi)
        
        
        # Instantiating widget properties 
        timeSlider_Label.setText("Time slider")
        ppm_Label.setText("Pixel per meter (ppm)")
        chunkSize_Label.setText("Chunk size (seconds)")
        speed_Label.setText("Speed filter (optional)")
        window_Label.setText("Window type")
        session_Label.setText("Current session")
        self.frequencyViewer_Label.setText("Frequency map")
        self.graph_Label.setText("Power spectrum graph")
        self.tracking_Label.setText("Animal tracking")
        self.bar.setOrientation(Qt.Vertical)
        # Start in green to signal no data rendered yet
        self.render_button.setStyleSheet("background-color : rgb(0, 180,0)")
        windowTypeBox.addItem("hann")
        windowTypeBox.addItem("hamming")
        windowTypeBox.addItem("blackmanharris")
        windowTypeBox.addItem("boxcar")
        speedTextBox.setPlaceholderText("Ex: Type 5,10 for 5cms to 10cms range filter")
        self.ppmTextBox.setText("511")
        chunkSizeTextBox.setText("60")
        
        # Set font sizes of label headers
        self.frequencyViewer_Label.setFont(QFont("Times New Roman", 18))
        self.graph_Label.setFont(QFont("Times New Roman", 18))
        self.tracking_Label.setFont(QFont("Times New Roman", 18))

        # Set header alignments
        self.frequencyViewer_Label.setAlignment(Qt.AlignCenter)
        self.graph_Label.setAlignment(Qt.AlignCenter)
        self.tracking_Label.setAlignment(Qt.AlignCenter)
        
        # Set fixed width and height for power display to fit all bands on single lines
        self.power_Label.setFixedWidth(380)
        self.power_Label.setFixedHeight(220)
        
        # Resize widgets to fixed width
        resizeWidgets = [windowTypeBox, chunkSizeTextBox, speedTextBox, self.ppmTextBox, 
                         self.browse_button, browse_folder_button]
        for widget in resizeWidgets:
            widget.setFixedWidth(300)
        
        # Set width of righthand buttons
        quit_button.setFixedWidth(150)
        save_button.setFixedWidth(150)
        self.render_button.setFixedWidth(150)
        self.load_eoi_btn.setFixedWidth(150)

        # Placing widgets
        # Top row buttons in order: Browse files, Browse folder, Load EOIs, Re-Render, Save data, Quit
        self.layout.addWidget(self.browse_button, 0, 0)
        self.layout.addWidget(browse_folder_button, 0, 1)
        self.layout.addWidget(self.load_eoi_btn, 0, 2)
        self.layout.addWidget(self.render_button, 0, 3)
        self.layout.addWidget(save_button, 0, 4)
        self.layout.addWidget(quit_button, 0, 5)
        
        # Second row
        self.layout.addWidget(session_Label, 1, 0)
        self.layout.addWidget(self.session_Text, 1, 1, 1, 2)  # Span 2 columns
        
        # Settings rows
        self.layout.addWidget(window_Label, 2, 0)
        self.layout.addWidget(windowTypeBox, 2, 1)
        self.layout.addWidget(ppm_Label, 3, 0)
        self.layout.addWidget(self.ppmTextBox, 3, 1)
        self.layout.addWidget(chunkSize_Label, 4, 0)
        self.layout.addWidget(chunkSizeTextBox, 4, 1)
        self.layout.addWidget(speed_Label, 5, 0)
        self.layout.addWidget(speedTextBox, 5, 1)
        self.layout.addWidget(self.graph_mode_button, 6, 0)
        self.layout.addWidget(self.frequencyViewer_Label, 6, 1)
        self.layout.addWidget(self.graph_Label, 6, 1)
        self.layout.addWidget(self.tracking_Label, 6, 3)
        self.layout.addWidget(self.power_Label, 7, 0)
        # Span frequency map/graph across columns 1-2 and tracking across columns 3-4
        self.layout.addWidget(self.view, 7, 1, 1, 2)
        self.layout.addWidget(self.graph_canvas, 7, 1, 1, 2)
        self.layout.addWidget(self.tracking_canvas, 7, 3, 1, 2)
        self.layout.addWidget(self.bar, 7, 5)
        # Place binned analysis button near the bottom-right
        self.layout.addWidget(self.binned_analysis_btn, 9, 4, alignment=Qt.AlignRight)
        self.layout.addWidget(timeSlider_Label, 8, 0)
        
        slider_layout = QHBoxLayout()
        slider_layout.addWidget(self.slider)
        slider_layout.addWidget(self.full_duration_btn)
        self.layout.addLayout(slider_layout, 8, 1, 1, 2)
        
        self.layout.addWidget(self.timeInterval_Label, 8, 3, 1, 2)
        self.layout.addWidget(self.progressBar_Label, 9, 5)
        self.layout.setSpacing(10)
        
        # Set column stretch to ensure frequency map and tracking plots resize equally
        # Frequency map spans columns 1-2, tracking spans columns 3-4
        self.layout.setColumnStretch(1, 1)
        self.layout.setColumnStretch(2, 1)
        self.layout.setColumnStretch(3, 1)
        self.layout.setColumnStretch(4, 1)
        
        # Hiding the canvas and graph label widget on startup 
        self.graph_canvas.close()
        self.graph_Label.close()
        
        # Widget signaling
        self.ppmTextBox.textChanged[str].connect(partial(self.textBoxChanged, 'ppm'))
        chunkSizeTextBox.textChanged[str].connect(partial(self.textBoxChanged, 'chunk_size'))
        speedTextBox.textChanged[str].connect(partial(self.textBoxChanged, 'speed'))
        quit_button.clicked.connect(self.quitClicked)
        self.browse_button.clicked.connect(self.runSession)
        browse_folder_button.clicked.connect(self.browseFolderClicked)
        self.graph_mode_button.clicked.connect(self.switch_graph)
        save_button.clicked.connect(self.saveClicked)
        self.render_button.clicked.connect(self.runSession)
        self.load_eoi_btn.clicked.connect(self.loadEOIFile)
        self.binned_analysis_btn.clicked.connect(self.openBinnedAnalysisWindow)
        self.slider.valueChanged[int].connect(self.sliderChanged)
        windowTypeBox.activated[str].connect(self.windowChanged)

        # Color the Browse button red if no EEG/EGF is loaded yet
        self.updateBrowseButtonStyle()
        
        # Auto-load EEG file if passed from command line
        if self.eeg_file_arg:
            QTimer.singleShot(100, self.loadFileFromArgument)
        
    # ------------------------------------------- #  
    
    def loadFileFromArgument(self):
        '''
            Auto-load the EEG/EGF file passed from STLAR main window.
            Also set PPM if provided.
        '''
        if not self.eeg_file_arg or not os.path.exists(self.eeg_file_arg):
            return
        
        try:
            # Set the active folder
            self.active_folder = os.path.dirname(os.path.realpath(self.eeg_file_arg))
            self.files = [None, None]
            
            # Determine file type and set accordingly
            ext = os.path.splitext(self.eeg_file_arg)[1].lower()
            if ('eeg' in ext) or ('egf' in ext):
                self.files[1] = self.eeg_file_arg
            
            # Try to find matching .pos file
            if self.files[1]:
                base_no_ext = os.path.splitext(self.files[1])[0]
                candidate_pos = base_no_ext + '.pos'
                candidate_pos_upper = base_no_ext + '.POS'
                if os.path.exists(candidate_pos):
                    self.files[0] = candidate_pos
                elif os.path.exists(candidate_pos_upper):
                    self.files[0] = candidate_pos_upper
                else:
                    # Try glob for any .pos sharing the same base name segment before last dot
                    base_dir = os.path.dirname(self.files[1])
                    base_name = os.path.basename(base_no_ext)
                    # Find any file that starts with base_name and ends with .pos
                    for name in os.listdir(base_dir):
                        if name.lower().endswith('.pos') and name.startswith(base_name):
                            self.files[0] = os.path.join(base_dir, name)
                            break
            
            # If we have both files, set session text
            if self.files[1] and self.files[0]:
                self.session_Text.setText(str(self.files[1]))
                self.updateBrowseButtonStyle()
            
            # Set PPM if provided
            if self.ppm_arg is not None:
                try:
                    ppm_value = float(self.ppm_arg)
                    self.ppm = ppm_value
                    # Update the PPM text box (block signals to avoid triggering textBoxChanged)
                    self.ppmTextBox.blockSignals(True)
                    self.ppmTextBox.setText(str(int(ppm_value)))
                    self.ppmTextBox.blockSignals(False)
                except (ValueError, TypeError):
                    pass
        except Exception as e:
            import traceback
            print(f"Error loading file from argument: {e}")
            traceback.print_exc()
    
    # ------------------------------------------- #  
    
    def textBoxChanged(self, label):
        
        '''
            Invoked when any one of the textboxes have their text changed.
            Handles the new input, sets variables.

            Params: 
                label (str) : 
                    The textbox label as a string. Chose between 'Speed', 
                    'ppm', and 'chunk_size'

            Returns: No return
        '''

        cbutton = self.sender()
        # Safety check: ensure sender has text() method
        if not hasattr(cbutton, 'text'):
            return
        # If any parameter is changed, will highlight
        # the re-render button to signal a re-rendering is needed
        curr_string = str(cbutton.text()).split(',')
        self.render_button.setStyleSheet("background-color : rgb(0, 180,0)")
        
        # The following fields are only set if field inputs are numeric
        # Sets speed filtering limits 
        if label == 'speed':
            if len(curr_string) == 1 and len(curr_string[0]) == 0:
                self.speed_lowerbound = self.speed_upperbound = None
            elif len(curr_string) == 1:
                if curr_string[0].isnumeric():
                    self.speed_lowerbound = int(curr_string[0])
                    self.speed_upperbound = 100
            elif len(curr_string) == 2:
                if curr_string[0].isnumeric() and curr_string[1].isnumeric():
                    self.speed_lowerbound = int(curr_string[0])
                    self.speed_upperbound = int(curr_string[1])
        # Sets ppm           
        elif label == 'ppm': 
            if len(curr_string) == 1 and len(curr_string[0]) == 0:
                self.ppm = None
            if curr_string[0].isnumeric():
                self.ppm = int(curr_string[0])
        # Sets chunk size        
        elif label == 'chunk_size':
            if len(curr_string) == 1 and len(curr_string[0]) == 0:
                self.chunk_size = None
            if curr_string[0].isnumeric():
                self.chunk_size = int(curr_string[0])
        
     # ------------------------------------------- # 
     
    def openFileNamesDialog(self):
    
        '''
            Will query OS to open file dialog box to choose pos and eeg/egf file.
            Additionally remembers the last opened directory. 
            
            Returns: 
                bool: bool value depending on whether appropriate files were chosen.
        '''

        # Set file dialog options
        options = QFileDialog.Options()
        # Open file dialog with file filter (include numbered variants)
        file_filter = (
            "EEG/EGF Files (*.eeg *.eeg2 *.eeg3 *.eeg4 *.egf *.egf2 *.egf3 *.egf4 *.pos);;"
            "EEG Files (*.eeg *.eeg2 *.eeg3 *.eeg4);;"
            "EGF Files (*.egf *.egf2 *.egf3 *.egf4);;"
            "Position Files (*.pos);;"
            "All Files (*)"
        )
        files, _ = QFileDialog.getOpenFileNames(self, "Choose .eeg/.egf file", self.active_folder, file_filter, options=options)
        if len(files) > 0:
            self.active_folder = dir_path = os.path.dirname(os.path.realpath((files[0])))
        else:
            return False

        # Reset tracked selection
        self.files = [None, None]

        # Helper to set by extension (case-insensitive)
        for file in files:
            ext = os.path.splitext(file)[1].lower()
            if ('pos' in ext):
                self.files[0] = file
            elif ('eeg' in ext) or ('egf' in ext):
                self.files[1] = file

        # If only EEG/EGF was selected, try auto-find matching .pos by basename
        if self.files[1] and not self.files[0]:
            eeg_path = self.files[1]
            base_no_ext = os.path.splitext(eeg_path)[0]
            candidate_pos = base_no_ext + '.pos'
            candidate_pos_upper = base_no_ext + '.POS'
            if os.path.exists(candidate_pos):
                self.files[0] = candidate_pos
            elif os.path.exists(candidate_pos_upper):
                self.files[0] = candidate_pos_upper
            else:
                # Try glob for any .pos sharing the same base name segment before last dot
                base_dir = os.path.dirname(eeg_path)
                base_name = os.path.basename(base_no_ext)
                # Find any file that starts with base_name and ends with .pos
                for name in os.listdir(base_dir):
                    if name.lower().endswith('.pos') and name.startswith(base_name):
                        self.files[0] = os.path.join(base_dir, name)
                        break

        # Validate we have at least EEG/EGF; and ensure POS exists (either selected or auto-found)
        if not self.files[1]:
            self.error_dialog.showMessage('Please select an .eeg/.egf file.')
            return False
        if not self.files[0]:
            self.error_dialog.showMessage('Matching .pos file not found. Please select the .pos file as well.')
            return False

        # Reflect session name using the electrophysiology file
        self.session_Text.setText(str(self.files[1]))
        self.updateBrowseButtonStyle()
        return True
    # ------------------------------------------- #
    
    def updatePowerDisplay(self, chunk_index):
        '''
            Update the power label to show all frequency bands and their 
            percentages for the current time chunk.
        '''
        if self.chunk_powers_data is None:
            return
        
        # Define frequency bands in order
        freq_bands = ['Delta', 'Theta', 'Beta', 'Low Gamma', 'High Gamma', 'Ripple', 'Fast Ripple']
        
        # Calculate total power across all bands for this chunk
        total_power = 0
        band_powers = {}
        for band in freq_bands:
            if band in self.chunk_powers_data and chunk_index < len(self.chunk_powers_data[band]):
                band_powers[band] = self.chunk_powers_data[band][chunk_index][0]
                total_power += band_powers[band]
        
        # Build the display text with all bands and their percentages
        if total_power > 0:
            display_text = "<b>Power Distribution:</b><br>"
            for band in freq_bands:
                if band in band_powers:
                    percentage = (band_powers[band] / total_power) * 100
                    # Highlight the currently selected band
                    if band == self.frequencyBand:
                        display_text += f"<b>\u2192 {band}: {percentage:.2f}%</b><br>"
                    else:
                        display_text += f"&nbsp;&nbsp;&nbsp;{band}: {percentage:.2f}%<br>"
            self.power_Label.setHtml(display_text)
        else:
            self.power_Label.setPlainText("No data available")

    # ------------------------------------------- #

    def updateBrowseButtonStyle(self):
        '''Set Browse button red when no EEG/EGF is selected.'''
        if getattr(self, 'browse_button', None) is None:
            return
        if self.files[1]:
            # Restore default look when session has EEG/EGF
            self.browse_button.setStyleSheet(self.browse_button_default_style)
        else:
            # Red background to signal missing electrophysiology file
            self.browse_button.setStyleSheet("background-color : rgb(200, 0, 0); color: white")
    
    # ------------------------------------------- #
    
    def quitClicked(self):

        '''
            Application exit
        '''

        print('quit')
        QApplication.quit()
        self.close() 
    
    # ------------------------------------------- #
    
    def browseFolderClicked(self):
        
        '''
            Opens folder selection dialog and initiates batch processing
            of all EGF/EEG files in the selected folder.
        '''
        
        # Prepare error dialog window 
        self.error_dialog = QErrorMessage()
        
        # Error checking ppm 
        if self.ppm == None or self.chunk_size == None: 
            self.error_dialog.showMessage('PPM field and/or Chunk Size field is blank, or has a non-numeric input. Please enter appropriate numerics.')
            return
        
        # If speed input only specifies lower bound, set upperbound to default
        if (self.speed_lowerbound != None and self.speed_upperbound == None):
            self.speed_upperbound = 100
            
        # If speed filter text is left blank, set default to 0cms to 100cms
        if self.speed_lowerbound == None and self.speed_upperbound == None: 
            self.speed_lowerbound = 0
            self.speed_upperbound = 100
        
        # Check speed bounds are ascending
        if self.speed_lowerbound != None and self.speed_upperbound != None:
            if self.speed_lowerbound > self.speed_upperbound: 
                self.error_dialog.showMessage('Speed filter range must be ascending. Lower speed first, higher speed after. Ex: 1,5')
                return
        
        # Open folder dialog
        options = QFileDialog.Options()
        options |= QFileDialog.ShowDirsOnly
        folder_path = QFileDialog.getExistingDirectory(
            self, 
            "Select folder containing EGF/EEG files", 
            self.active_folder,
            options=options
        )
        
        if not folder_path:
            return
        
        # Remember the selected folder
        self.active_folder = folder_path
        
        # Update session label
        self.session_Text.setText(f"Batch: {folder_path}")
        self.updateBrowseButtonStyle()
        
        # Disable buttons during processing
        self.setButtonsEnabled(False)
        
        # Create and start batch worker thread
        self.batch_worker = BatchWorkerThread(
            folder_path,
            self.ppm,
            self.chunk_size,
            self.window_type,
            self.speed_lowerbound,
            self.speed_upperbound,
            output_dir=folder_path
        )
        
        # Connect signals
        self.batch_worker.progress_update.connect(self.updateBatchLabel)
        self.batch_worker.progress_value.connect(self.progressBar)
        self.batch_worker.finished_signal.connect(self.batchProcessingFinished)
        self.batch_worker.error_signal.connect(self.batchProcessingError)
        
        # Start processing
        self.batch_worker.start()
        self.progressBar_Label.setText("Batch processing in progress...")
    
    # ------------------------------------------- #
    
    def updateBatchLabel(self, text):
        '''Update progress label during batch processing'''
        self.progressBar_Label.setText(text)
        print(text)
    
    # ------------------------------------------- #
    
    def batchProcessingFinished(self, results):
        '''Handle completion of batch processing'''
        
        successful = results.get('successful', [])
        failed = results.get('failed', [])
        total = results.get('total', 0)
        
        # Re-enable buttons
        self.setButtonsEnabled(True)
        
        # Build summary message
        summary = f"Batch processing complete!\n\n"
        summary += f"Total files: {total}\n"
        summary += f"✓ Successful: {len(successful)}\n"
        
        if successful:
            summary += "\nProcessed files:\n"
            for fname in successful[:5]:  # Show first 5
                summary += f"  • {fname}\n"
            if len(successful) > 5:
                summary += f"  ... and {len(successful)-5} more"
        
        if failed:
            summary += f"\n✗ Failed: {len(failed)}\n"
            for fname, error in failed[:3]:  # Show first 3 errors
                summary += f"  • {fname}: {error[:50]}...\n"
            if len(failed) > 3:
                summary += f"  ... and {len(failed)-3} more"
        
        summary += f"\nCSV files saved to:\n{self.active_folder}"
        
        # Show summary dialog
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Batch Processing Complete")
        msg_box.setText(summary)
        msg_box.setIcon(QMessageBox.Information)
        msg_box.exec_()
        
        self.progressBar_Label.setText("Batch processing complete!")
        self.bar.setValue(100)
    
    # ------------------------------------------- #
    
    def batchProcessingError(self, error_msg):
        '''Handle errors during batch processing'''
        self.setButtonsEnabled(True)
        self.error_dialog = QErrorMessage()
        self.error_dialog.showMessage(f"Batch processing error: {error_msg}")
        self.progressBar_Label.setText("Batch processing error!")
    
    # ------------------------------------------- #
    
    def setButtonsEnabled(self, enabled):
        '''Enable/disable buttons during processing'''
        # Find all buttons and disable them
        for widget in self.findChildren(QPushButton):
            if widget.text() not in ['Quit']:  # Keep Quit button enabled
                widget.setEnabled(enabled)
    
    # ------------------------------------------- #
    
    def saveClicked(self):
        
        '''
            Automatically saves Excel and CSV with average frequency powers vs time
            into the output folder with filename suffix '_SSM'.
        '''

        # If there are no chunk powers, do nothing
        if self.chunk_powers_data is None:
            return

        # Try to use Excel; fallback to CSV
        try:
            import openpyxl
            use_excel = True
        except ImportError:
            use_excel = False
            import csv

        # Expected band order and labels
        bands = [
            ("Delta", "Avg Delta Power"),
            ("Theta", "Avg Theta Power"),
            ("Beta", "Avg Beta Power"),
            ("Low Gamma", "Avg Low Gamma Power"),
            ("High Gamma", "Avg High Gamma Power"),
            ("Ripple", "Avg Ripple Power"),
            ("Fast Ripple", "Avg Fast Ripple Power"),
        ]

        # Determine output directory and base name
        out_dir = self.active_folder or os.getcwd()
        base_name = os.path.splitext(os.path.basename(self.files[1] or 'output'))[0]
        
        if use_excel:
            out_file = os.path.join(out_dir, f"{base_name}_SSM.xlsx")
        else:
            out_file = os.path.join(out_dir, f"{base_name}_SSM.csv")

        # Build header and rows and write CSV or Excel
        band_labels = [label for _, label in bands]
        percent_labels = [f"Percent {name}" for name, _ in bands]
        
        # Calculate distances for each chunk if tracking data is available
        distances_per_bin = []
        cumulative_distances = []
        cumulative_sum = 0.0
        
        if self.tracking_data is not None:
            # Handle both 2-tuple and 3-tuple
            pos_x_chunks = self.tracking_data[0]
            pos_y_chunks = self.tracking_data[1]
            
            for i in range(len(pos_x_chunks)):
                distance_cm_in_bin = 0.0
                
                # Get the positions for this bin only
                x_bin = pos_x_chunks[i]
                y_bin = pos_y_chunks[i]
                
                # Calculate distance within this specific bin only
                if len(x_bin) > 1:
                    dx = np.diff(np.array(x_bin))
                    dy = np.diff(np.array(y_bin))
                    distances_in_bin_pixels = np.sqrt(dx**2 + dy**2)
                    total_distance_pixels_in_bin = np.sum(distances_in_bin_pixels)
                    
                    # Convert from pixels to centimeters using PPM
                    if self.ppm is not None and self.ppm > 0:
                        distance_cm_in_bin = (total_distance_pixels_in_bin / self.ppm) * 100
                    else:
                        distance_cm_in_bin = total_distance_pixels_in_bin
                
                # Store distance for this bin only (not cumulative)
                distances_per_bin.append(distance_cm_in_bin)
                
                # Update cumulative distance
                cumulative_sum += distance_cm_in_bin
                cumulative_distances.append(cumulative_sum)
        
        header = ["Time Bin (s)", "Distance Per Bin (cm)", "Cumulative Distance (cm)"] + band_labels + percent_labels
        
        # Check for EOIs
        has_eois = hasattr(self, 'eoi_segments') and self.eoi_segments
        if has_eois:
            header.append("EOI Count")
        
        # Determine chunk size for proper time bin labeling
        chunk_size = self.chunk_size if self.chunk_size is not None else 10
        
        # Calculate the actual recording duration and round down to nearest chunk boundary
        # This handles cases where recording is slightly over (e.g., 901s, 1201s, 601s)
        if len(self.pos_t) > 0:
            actual_duration = float(self.pos_t[-1])
            # Round down to nearest chunk boundary
            max_full_chunks = int(actual_duration / chunk_size)
            num_rows = min(len(self.pos_t), max_full_chunks)
        else:
            num_rows = len(self.pos_t)
        
        # Gather per-band arrays, ensure 1D length matches timestamps
        band_arrays = {}
        for key, label in bands:
            arr = np.array(self.chunk_powers_data.get(key, [])).reshape(-1)
            band_arrays[label] = arr
        
        # Prepare data rows
        data_rows = []
        for i in range(num_rows):
            # Create time bin string based on chunk size
            # Bins start at chunk_size because first chunk (0 to chunk_size) is used for baseline
            time_bin_start = i * chunk_size
            time_bin_end = (i + 1) * chunk_size
            time_bin_str = f"{time_bin_start}-{time_bin_end}"
            
            row = [time_bin_str]
            # Add distance per bin
            if distances_per_bin and i < len(distances_per_bin):
                row.append(round(distances_per_bin[i], 3))
            else:
                row.append("")
            # Add cumulative distance
            if cumulative_distances and i < len(cumulative_distances):
                row.append(round(cumulative_distances[i], 3))
            else:
                row.append("")
            band_values = []
            for _, label in bands:
                val = band_arrays[label][i] if i < len(band_arrays[label]) else ""
                band_values.append(val)
            # Append band values
            row.extend(band_values)
            # Compute total power across bands at this row (ignore blanks)
            numeric_vals = [float(v) for v in band_values if v != ""]
            total_power = sum(numeric_vals)
            # Compute per-band percentages
            for v in band_values:
                if v == "" or total_power == 0:
                    row.append("")
                else:
                    row.append(round((float(v) / total_power) * 100.0, 3))
            
            if has_eois:
                count = 0
                if i in self.eoi_segments:
                    count = len(self.eoi_segments[i])
                row.append(count)
            
            data_rows.append(row)
        
        # Write Excel or CSV
        if use_excel:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'SSM Data'
            ws.append(header)
            for row in data_rows:
                ws.append(row)
            wb.save(out_file)
        else:
            with open(out_file, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(header)
                for row in data_rows:
                    writer.writerow(row)
        
        # Save frequency map for entire duration (if available)
        saved_files = [out_file]
        if self.images is not None and len(self.images) > 0:
            try:
                freq_map_path = os.path.join(out_dir, f"{base_name}_FrequencyMap_FullDuration.png")
                avg_pixmap = self._calculate_average_pixmap(self.images)
                
                # Convert QPixmap to QImage and scale to double size for better quality
                qimage = avg_pixmap.toImage()
                original_width = qimage.width()
                original_height = qimage.height()
                scaled_image = qimage.scaled(original_width * 2, original_height * 2, 
                                             Qt.KeepAspectRatio, Qt.SmoothTransformation)
                scaled_pixmap = QPixmap.fromImage(scaled_image)
                scaled_pixmap.save(freq_map_path)
                saved_files.append(freq_map_path)
            except Exception as e:
                print(f"Could not save frequency map: {e}")
        
        # Save animal tracking for entire duration (if available)
        if self.tracking_data is not None:
            try:
                import matplotlib.pyplot as plt
                from matplotlib.figure import Figure
                
                tracking_path = os.path.join(out_dir, f"{base_name}_AnimalTracking_FullDuration.png")
                
                # Create a new figure for saving (larger size for better quality)
                fig = Figure(figsize=(12, 12), dpi=200)
                ax = fig.add_subplot(111)
                ax.set_xlabel('X - coordinates', fontsize=12)
                ax.set_ylabel('Y - coordinates', fontsize=12)
                ax.set_title('Animal Tracking - Full Duration', fontsize=14, fontweight='bold')
                
                # Set fixed limits and aspect ratio
                if hasattr(self, 'track_xlim'):
                    ax.set_xlim(self.track_xlim)
                    ax.set_ylim(self.track_ylim)
                    ax.set_aspect('equal', adjustable='box')
                    
                    # Draw bins (same as GUI)
                    self._draw_tracking_bins(ax)
                
                # Plot all position data (entire duration)
                x_all = [item for sublist in self.tracking_data[0] for item in sublist]
                y_all = [item for sublist in self.tracking_data[1] for item in sublist]
                ax.plot(x_all, y_all, linewidth=0.8, color='#D3D3D3', alpha=0.6, label='Path')
                
                # Plot all EOIs if available
                if hasattr(self, 'eoi_segments') and self.eoi_segments:
                    eoi_count = 0
                    for chunk_idx, segments in self.eoi_segments.items():
                        for seg_x, seg_y in segments:
                            if eoi_count == 0:
                                ax.plot(seg_x, seg_y, color='red', linewidth=2, alpha=0.8, label='EOIs')
                            else:
                                ax.plot(seg_x, seg_y, color='red', linewidth=2, alpha=0.8)
                            eoi_count += 1
                    ax.legend(loc='upper right', fontsize=10)
                
                # Add occupancy labels (same as GUI)
                occ_info = self._compute_tracking_occupancy(is_full=True, chunk_idx=0)
                if occ_info:
                    # Show occupancy if no EOIs or EOIs sum to zero
                    show_occ = True
                    if hasattr(self, 'eoi_segments') and self.eoi_segments:
                        eoi_info = self._compute_tracking_eoi_counts(is_full=True, chunk_idx=0)
                        if eoi_info:
                            vals = eoi_info.get('values')
                            try:
                                show_occ = (vals is None) or (np.sum(vals) <= 0)
                            except Exception:
                                show_occ = True
                    
                    if show_occ:
                        self._add_tracking_occupancy_labels(ax, occ_info)
                
                fig.savefig(tracking_path, format='png', dpi=200, bbox_inches='tight')
                plt.close(fig)
                saved_files.append(tracking_path)
            except Exception as e:
                print(f"Could not save animal tracking: {e}")
                import traceback
                traceback.print_exc()
        
        # Show success message with all saved files
        files_list = "\n".join([os.path.basename(f) for f in saved_files])
        format_str = "Excel" if use_excel else "CSV (openpyxl not installed)"
        QMessageBox.information(self, "Save Complete", 
            f"Saved files:\n\n{files_list}\n\nLocation: {out_dir}\nFormat: {format_str}")
            
    # ------------------------------------------- #
    
    def switch_graph(self):
        
        '''
            Will switch between plotting power spectral density per chunk to frequency map per chunk.
        '''

        cbutton = self.sender()
        # Show the PSD graph is graph mode is chosen
        if cbutton.text() == 'Graph mode':
            self.graph_canvas.show()
            self.view.close()
            self.frequencyViewer_Label.close()
            self.graph_Label.show()
            self.plot_flag = True
            self.graph_mode_button.setText("Frequency image mode")
            # Only show the graph if there is something to plot
            if self.chunk_index != None and self.plot_data is not None:
                freq, pdf = self.plot_data[self.chunk_index][0]
                
                # Plot with modern styling
                self.graph_canvas.axes.plot(freq, pdf, linewidth=2, color='#2E86AB', alpha=0.9)
                self.graph_canvas.axes.fill_between(freq, pdf, alpha=0.3, color='#2E86AB')
                
                # Add frequency band shading with ranges in labels
                freq_bands = {
                    'Delta (1-3 Hz)': (1, 3, '#FF6B6B'),
                    'Theta (4-12 Hz)': (4, 12, '#4ECDC4'),
                    'Beta (13-20 Hz)': (13, 20, '#95E1D3'),
                    'Low Gamma (35-55 Hz)': (35, 55, '#F38181'),
                    'High Gamma (65-120 Hz)': (65, 120, '#AA96DA'),
                    'Ripple (80-250 Hz)': (80, 250, '#FCBAD3'),
                    'Fast Ripple (250-500 Hz)': (250, 500, '#A8D8EA')
                }
                for band_label, (low, high, color) in freq_bands.items():
                    self.graph_canvas.axes.axvspan(low, high, alpha=0.1, color=color, label=band_label)
                
                # Styling
                self.graph_canvas.axes.set_xlabel('Frequency (Hz)', fontsize=11, fontweight='bold')
                self.graph_canvas.axes.set_ylabel('Power Spectral Density (µV²/Hz)', fontsize=11, fontweight='bold')
                self.graph_canvas.axes.grid(True, alpha=0.3, linestyle='--', linewidth=0.5)
                self.graph_canvas.axes.set_facecolor('#F8F9FA')
                self.graph_canvas.axes.spines['top'].set_visible(False)
                self.graph_canvas.axes.spines['right'].set_visible(False)
                self.graph_canvas.axes.legend(loc='upper right', fontsize=7, framealpha=0.9)
        # Else show freq map
        else:
            self.view.show()
            self.graph_canvas.close()
            self.graph_Label.close()
            self.frequencyViewer_Label.show()
            self.plot_flag = False
            self.graph_mode_button.setText("Graph mode")
        
    # ------------------------------------------- #  
    
    def updatePowerDisplay(self, chunk_index):
        '''
            Update the power label to show all frequency bands with ranges and their 
            percentages for the current time chunk.
        '''
        if self.chunk_powers_data is None:
            return
        
        # Define frequency bands with ranges
        freq_band_ranges = {
            'Delta': '1-3 Hz',
            'Theta': '4-12 Hz',
            'Beta': '13-20 Hz',
            'Low Gamma': '35-55 Hz',
            'High Gamma': '65-120 Hz',
            'Ripple': '80-250 Hz',
            'Fast Ripple': '250-500 Hz'
        }
        freq_bands = list(freq_band_ranges.keys())
        
        # Calculate total power across all bands for this chunk
        total_power = 0
        band_powers = {}
        for band in freq_bands:
            if band in self.chunk_powers_data and chunk_index < len(self.chunk_powers_data[band]):
                band_powers[band] = self.chunk_powers_data[band][chunk_index][0]
                total_power += band_powers[band]
        
        # Build the display text with all bands, ranges, and their percentages
        if total_power > 0:
            display_text = "<b>Power Distribution:</b><br>"
            for band in freq_bands:
                if band in band_powers:
                    percentage = (band_powers[band] / total_power) * 100
                    freq_range = freq_band_ranges[band]
                    display_text += f"{band} ({freq_range}): {percentage:.2f}%<br>"
            self.power_Label.setHtml(display_text)
        else:
            self.power_Label.setPlainText("No data available")
    
    # ------------------------------------------- #
    
    def updatePowerDisplayFull(self):
        '''
            Update the power label to show average power distribution over the full duration.
        '''
        if self.chunk_powers_data is None:
            return
        
        freq_band_ranges = {
            'Delta': '1-3 Hz', 'Theta': '4-12 Hz', 'Beta': '13-20 Hz',
            'Low Gamma': '35-55 Hz', 'High Gamma': '65-120 Hz',
            'Ripple': '80-250 Hz', 'Fast Ripple': '250-500 Hz'
        }
        freq_bands = list(freq_band_ranges.keys())
        
        total_power = 0
        band_powers = {}
        
        for band in freq_bands:
            if band in self.chunk_powers_data:
                vals = [x[0] for x in self.chunk_powers_data[band]]
                s = sum(vals)
                band_powers[band] = s
                total_power += s
        
        if total_power > 0:
            display_text = "<b>Full Duration Power Distribution:</b><br>"
            for band in freq_bands:
                if band in band_powers:
                    percentage = (band_powers[band] / total_power) * 100
                    freq_range = freq_band_ranges[band]
                    display_text += f"{band} ({freq_range}): {percentage:.2f}%<br>"
            self.power_Label.setHtml(display_text)
        else:
            self.power_Label.setPlainText("No data available")

    # ------------------------------------------- #
    
    def sliderChanged(self, value): 
        
        '''
            Create a slider that allows the user to sift through each chunk and
            view how the graph/frequency maps change as a function of time.
            Also handles Full Duration view.
        '''

        # Sliders value acts as chunk index
        self.chunk_index = value 
        
        is_full = hasattr(self, 'full_duration_btn') and self.full_duration_btn.isChecked()
        
        # If we have data to plot, plot the graph
        if self.plot_flag:
            if self.plot_data is not None:
                self.graph_canvas.axes.cla()
                
                if is_full:
                    # Average PSD across all chunks
                    all_pdfs = []
                    freq = None
                    for chunk in self.plot_data:
                        f, p = chunk[0]
                        if freq is None: freq = f
                        all_pdfs.append(p)
                    
                    if all_pdfs:
                        avg_pdf = np.mean(all_pdfs, axis=0)
                        self.graph_canvas.axes.plot(freq, avg_pdf, linewidth=2, color='#2E86AB', alpha=0.9)
                        self.graph_canvas.axes.fill_between(freq, avg_pdf, alpha=0.3, color='#2E86AB')
                        self.graph_canvas.axes.set_title("Average Power Spectral Density (Full Duration)", fontsize=10)
                else:
                    freq, pdf = self.plot_data[value][0]
                    # Plot with modern styling
                    self.graph_canvas.axes.plot(freq, pdf, linewidth=2, color='#2E86AB', alpha=0.9)
                    self.graph_canvas.axes.fill_between(freq, pdf, alpha=0.3, color='#2E86AB')
                
                # Add frequency band shading
                freq_bands = {
                    'Delta': (1, 3, '#FF6B6B'),
                    'Theta': (4, 12, '#4ECDC4'),
                    'Beta': (13, 20, '#95E1D3'),
                    'Low Gamma': (35, 55, '#F38181'),
                    'High Gamma': (65, 120, '#AA96DA'),
                    'Ripple': (80, 250, '#FCBAD3'),
                    'Fast Ripple': (250, 500, '#A8D8EA')
                }
                # y_max = pdf.max() * 1.1
                for band_name, (low, high, color) in freq_bands.items():
                    self.graph_canvas.axes.axvspan(low, high, alpha=0.1, color=color, label=band_name)
                
                # Styling
                self.graph_canvas.axes.set_xlabel('Frequency (Hz)', fontsize=11, fontweight='bold')
                self.graph_canvas.axes.set_ylabel('Power Spectral Density (µV²/Hz)', fontsize=11, fontweight='bold')
                self.graph_canvas.axes.grid(True, alpha=0.3, linestyle='--', linewidth=0.5)
                self.graph_canvas.axes.set_facecolor('#F8F9FA')
                self.graph_canvas.axes.spines['top'].set_visible(False)
                self.graph_canvas.axes.spines['right'].set_visible(False)
                
                # Add legend
                self.graph_canvas.axes.legend(loc='upper right', fontsize=7, framealpha=0.9)
                
                self.graph_canvas.draw()
        # Else if we have maps to plot, plot the frequency maps    
        elif self.images is not None:
            if is_full:
                avg_pixmap = self._calculate_average_pixmap(self.images)
                self.imageMapper.setPixmap(avg_pixmap)
            else:
                self.imageMapper.setPixmap(self.images[value])
        
        if self.tracking_data is not None:
            self.tracking_canvas.axes.cla()
            self.tracking_canvas.axes.set_xlabel('X - coordinates')
            self.tracking_canvas.axes.set_ylabel('Y - coordinates')
            
            # Set fixed limits and aspect ratio to prevent jitter
            if hasattr(self, 'track_xlim'):
                self.tracking_canvas.axes.set_xlim(self.track_xlim)
                self.tracking_canvas.axes.set_ylim(self.track_ylim)
                self.tracking_canvas.axes.set_aspect('equal', adjustable='box')
                # Draw bins
                self._draw_tracking_bins(self.tracking_canvas.axes)
            
            # Precompute EOI info first so we can decide whether to show occupancy labels
            eoi_info = self._compute_tracking_eoi_counts(is_full, value)
            occ_info = self._compute_tracking_occupancy(is_full, value)
            
            if is_full:
                # Plot all chunks
                x_all = [item for sublist in self.tracking_data[0] for item in sublist]
                y_all = [item for sublist in self.tracking_data[1] for item in sublist]
                self.tracking_canvas.axes.plot(x_all, y_all, linewidth=0.5, color='#D3D3D3', alpha=0.6)
                self.tracking_canvas.axes.set_title("Full Path", fontsize=10)
                
                # Plot all EOIs
                if self.eoi_segments:
                    for chunk_idx, segments in self.eoi_segments.items():
                        for seg_x, seg_y in segments:
                            self.tracking_canvas.axes.plot(seg_x, seg_y, color='red', linewidth=1.5, alpha=0.8)
            else:
                self.tracking_canvas.axes.plot(self.tracking_data[0][value], self.tracking_data[1][value], linewidth=0.5, color='#D3D3D3')
                
                # Plot EOIs if available for this chunk
                if value in self.eoi_segments:
                    for seg_x, seg_y in self.eoi_segments[value]:
                        self.tracking_canvas.axes.plot(seg_x, seg_y, color='red', linewidth=1.5, alpha=0.8)
            
            # Only show occupancy labels when no EOIs are present to avoid overlap
            show_occ = False
            if occ_info:
                if not eoi_info:
                    show_occ = True
                else:
                    vals = eoi_info.get('values')
                    try:
                        show_occ = (vals is None) or (np.sum(vals) <= 0)
                    except Exception:
                        show_occ = False
            if show_occ:
                self._add_tracking_occupancy_labels(self.tracking_canvas.axes, occ_info)

            # Overlay EOI counts per bin (red), for chunk/full view
            if eoi_info:
                self._add_tracking_eoi_labels(self.tracking_canvas.axes, eoi_info)
            
            self.tracking_canvas.draw()
        # Reflect chunk and its nominal time range (consistent with binned slider)
        if self.pos_t is not None and self.chunk_size is not None:
            if is_full:
                total_duration = float(self.pos_t[-1])
                self.timeInterval_Label.setText(f"Full Duration (0-{total_duration:.0f}s)")
            else:
                total_chunks = (self.slider.maximum() + 1) if self.slider is not None else 1
                t_start = value * self.chunk_size
                # Clamp end to actual duration for the last chunk if binned data is available
                if hasattr(self, 'binned_data') and self.binned_data and 'duration' in self.binned_data:
                    duration = float(self.binned_data.get('duration', (total_chunks * self.chunk_size)))
                    t_end_nominal = (value + 1) * self.chunk_size
                    t_end = min(t_end_nominal, duration)
                else:
                    t_end = (value + 1) * self.chunk_size
                self.timeInterval_Label.setText(f"Chunk {value+1}/{total_chunks} ({t_start:.0f}-{t_end:.0f}s)")
        
        # Update power display with current chunk's frequency band percentages
        if is_full:
            self.updatePowerDisplayFull()
        else:
            self.updatePowerDisplay(value)

    # ------------------------------------------- #
    
    def toggleFullDuration(self):
        '''Toggle between chunk view and full duration view'''
        is_full = self.full_duration_btn.isChecked()
        self.slider.setEnabled(not is_full)
        self.sliderChanged(self.slider.value())

    def _draw_tracking_bins(self, ax):
        '''Draw spatial bins (polar or grid) on the tracking plot'''
        if not hasattr(self, 'data_bounds'): return
        min_x, max_x, min_y, max_y = self.data_bounds
        
        is_polar = False
        if hasattr(self, 'binned_data') and self.binned_data:
             is_polar = (self.binned_data.get('type') == 'polar')
        elif hasattr(self, 'arena_shape'):
             is_polar = ("Circle" in self.arena_shape or "Ellipse" in self.arena_shape)
             
        if is_polar:
            width = max_x - min_x
            height = max_y - min_y
            center_x = min_x + width/2
            center_y = min_y + height/2
            
            # Outer (r=1)
            e1 = Ellipse((center_x, center_y), width, height, fill=False, edgecolor='gray', linestyle='--', alpha=0.5)
            ax.add_patch(e1)
            # Inner (r=0.707)
            scale = 1.0 / np.sqrt(2)
            e2 = Ellipse((center_x, center_y), width*scale, height*scale, fill=False, edgecolor='gray', linestyle='--', alpha=0.5)
            ax.add_patch(e2)
            
            # Sectors
            angles = np.linspace(-np.pi, np.pi, 9)
            for theta in angles:
                x_edge = center_x + (width/2) * np.cos(theta)
                y_edge = center_y + (height/2) * np.sin(theta)
                ax.plot([center_x, x_edge], [center_y, y_edge], color='gray', linestyle='--', alpha=0.5, linewidth=0.8)
        else:
            # 4x4 Grid
            x_edges = np.linspace(min_x, max_x, 5)
            y_edges = np.linspace(min_y, max_y, 5)
            for x in x_edges:
                ax.axvline(x, color='gray', linestyle='--', alpha=0.5, linewidth=0.8)
            for y in y_edges:
                ax.axhline(y, color='gray', linestyle='--', alpha=0.5, linewidth=0.8)

    def _compute_tracking_occupancy(self, is_full, chunk_idx):
        '''Compute per-bin occupancy percentages for tracking plot.'''
        if self.tracking_data is None or not hasattr(self, 'data_bounds'):
            return None
        min_x, max_x, min_y, max_y = self.data_bounds
        is_polar = False
        if hasattr(self, 'binned_data') and self.binned_data:
            is_polar = (self.binned_data.get('type') == 'polar')
        elif hasattr(self, 'arena_shape'):
            is_polar = ("Circle" in self.arena_shape or "Ellipse" in self.arena_shape)
        if is_full:
            xs = np.concatenate(self.tracking_data[0]) if self.tracking_data[0] else np.array([])
            ys = np.concatenate(self.tracking_data[1]) if self.tracking_data[1] else np.array([])
        else:
            if chunk_idx >= len(self.tracking_data[0]):
                return None
            xs = np.array(self.tracking_data[0][chunk_idx])
            ys = np.array(self.tracking_data[1][chunk_idx])
        if xs.size == 0 or ys.size == 0:
            return None
        if is_polar:
            width = max_x - min_x
            height = max_y - min_y
            if width <= 0 or height <= 0:
                return None
            center_x = min_x + width/2
            center_y = min_y + height/2
            if self.binned_data and 'bin_occupancy_timeseries' in self.binned_data:
                occ_data = None
                if is_full:
                    occ_data = np.sum(self.binned_data['bin_occupancy_timeseries'], axis=2)
                else:
                    occ_ts = self.binned_data['bin_occupancy_timeseries']
                    if chunk_idx < occ_ts.shape[2]:
                        occ_data = occ_ts[:, :, chunk_idx]
                if occ_data is None:
                    return None
            else:
                dx = xs - center_x
                dy = ys - center_y
                rx = width / 2.0 if width > 0 else 1.0
                ry = height / 2.0 if height > 0 else 1.0
                r_norm = np.sqrt((dx / rx) ** 2 + (dy / ry) ** 2)
                theta = np.arctan2(dy, dx)
                r_edges = np.array([0.0, 1.0 / np.sqrt(2), 1e9])
                theta_edges = np.linspace(-np.pi, np.pi, 9)
                occ_data, _, _ = np.histogram2d(r_norm, theta, bins=[r_edges, theta_edges])
            total = np.sum(occ_data)
            if total <= 0:
                return None
            occ_pct = (occ_data / total) * 100.0
            return {
                'is_polar': True,
                'values': occ_pct,
                'radius_edges': np.array([0.0, 1.0 / np.sqrt(2), 1.0]),
                'angle_edges': np.linspace(-np.pi, np.pi, 9),
                'center': (center_x, center_y),
                'width': width,
                'height': height
            }
        x_edges = np.linspace(min_x, max_x, 5)
        y_edges = np.linspace(min_y, max_y, 5)
        occ_counts, _, _ = np.histogram2d(xs, ys, bins=[x_edges, y_edges])
        total = np.sum(occ_counts)
        if total <= 0:
            return None
        occ_pct = (occ_counts / total) * 100.0
        x_centers = 0.5 * (x_edges[:-1] + x_edges[1:])
        y_centers = 0.5 * (y_edges[:-1] + y_edges[1:])
        return {
            'is_polar': False,
            'values': occ_pct,
            'x_centers': x_centers,
            'y_centers': y_centers
        }

    def _add_tracking_occupancy_labels(self, ax, occ_info):
        '''Overlay occupancy percent labels on the tracking plot.'''
        if not occ_info or 'values' not in occ_info:
            return
        values = occ_info['values']
        if np.sum(values) <= 0:
            return
        if occ_info.get('is_polar'):
            center_x, center_y = occ_info['center']
            width = occ_info['width']
            height = occ_info['height']
            r_edges = occ_info['radius_edges']
            theta_edges = occ_info['angle_edges']
            r_centers = 0.5 * (r_edges[:-1] + r_edges[1:])
            theta_centers = 0.5 * (theta_edges[:-1] + theta_edges[1:])
            for i, r in enumerate(r_centers):
                for j, theta in enumerate(theta_centers):
                    pct = values[i, j]
                    if pct <= 0:
                        continue
                    r_clamped = min(r, 1.0)
                    x = center_x + (width / 2.0) * r_clamped * np.cos(theta)
                    y = center_y + (height / 2.0) * r_clamped * np.sin(theta)
                    ax.text(x, y, f"{pct:.0f}%", fontsize=7, ha='center', va='center', color='#222',
                            bbox=dict(boxstyle='round,pad=0.15', facecolor='white', alpha=0.65, linewidth=0))
        else:
            x_centers = occ_info['x_centers']
            y_centers = occ_info['y_centers']
            for i, xc in enumerate(x_centers):
                for j, yc in enumerate(y_centers):
                    pct = values[i, j]
                    if pct <= 0:
                        continue
                    ax.text(xc, yc, f"{pct:.0f}%", fontsize=7, ha='center', va='center', color='#222',
                            bbox=dict(boxstyle='round,pad=0.15', facecolor='white', alpha=0.65, linewidth=0))

    def _compute_tracking_eoi_counts(self, is_full, chunk_idx):
        '''Compute per-bin event counts (not position samples) for tracking plot.'''
        if not hasattr(self, 'data_bounds') or not hasattr(self, 'eoi_segments') or not self.eoi_segments:
            return None
        min_x, max_x, min_y, max_y = self.data_bounds
        is_polar = False
        if hasattr(self, 'binned_data') and self.binned_data:
            is_polar = (self.binned_data.get('type') == 'polar')
        elif hasattr(self, 'arena_shape'):
            is_polar = ("Circle" in self.arena_shape or "Ellipse" in self.arena_shape)

        # Gather EOI segments (each segment = one event)
        event_segments = []
        if is_full:
            for _, segments in self.eoi_segments.items():
                event_segments.extend(segments)
        else:
            if chunk_idx in self.eoi_segments:
                event_segments = self.eoi_segments[chunk_idx]

        if not event_segments:
            return None

        if is_polar:
            width = max_x - min_x
            height = max_y - min_y
            if width <= 0 or height <= 0:
                return None
            center_x = min_x + width/2
            center_y = min_y + height/2
            
            # Count events per bin (each event counted once in bins it touches)
            event_counts = np.zeros((2, 8))
            r_edges = np.array([0.0, 1.0 / np.sqrt(2), 1e9])
            theta_edges = np.linspace(-np.pi, np.pi, 9)
            
            for seg_x, seg_y in event_segments:
                if len(seg_x) == 0 or len(seg_y) == 0:
                    continue
                xs = np.asarray(seg_x)
                ys = np.asarray(seg_y)
                dx = xs - center_x
                dy = ys - center_y
                rx = width / 2.0 if width > 0 else 1.0
                ry = height / 2.0 if height > 0 else 1.0
                r_norm = np.sqrt((dx / rx) ** 2 + (dy / ry) ** 2)
                theta = np.arctan2(dy, dx)
                
                # Determine which bins this event touches
                r_indices = np.digitize(r_norm, r_edges) - 1
                r_indices = np.clip(r_indices, 0, 1)
                theta_indices = np.digitize(theta, theta_edges) - 1
                theta_indices = np.clip(theta_indices, 0, 7)
                
                # Count this event once per unique bin it touches
                unique_bins = set(zip(r_indices, theta_indices))
                for ri, ti in unique_bins:
                    event_counts[ri, ti] += 1
            
            return {
                'is_polar': True,
                'values': event_counts,
                'radius_edges': np.array([0.0, 1.0 / np.sqrt(2), 1.0]),
                'angle_edges': theta_edges,
                'center': (center_x, center_y),
                'width': width,
                'height': height
            }

        # Cartesian (4x4) - count events per bin
        x_edges = np.linspace(min_x, max_x, 5)
        y_edges = np.linspace(min_y, max_y, 5)
        event_counts = np.zeros((4, 4))
        
        for seg_x, seg_y in event_segments:
            if len(seg_x) == 0 or len(seg_y) == 0:
                continue
            xs = np.asarray(seg_x)
            ys = np.asarray(seg_y)
            
            # Determine which bins this event touches
            x_indices = np.digitize(xs, x_edges) - 1
            x_indices = np.clip(x_indices, 0, 3)
            y_indices = np.digitize(ys, y_edges) - 1
            y_indices = np.clip(y_indices, 0, 3)
            
            # Count this event once per unique bin it touches
            unique_bins = set(zip(x_indices, y_indices))
            for xi, yi in unique_bins:
                event_counts[xi, yi] += 1
        
        x_centers = 0.5 * (x_edges[:-1] + x_edges[1:])
        y_centers = 0.5 * (y_edges[:-1] + y_edges[1:])
        return {
            'is_polar': False,
            'values': event_counts,
            'x_centers': x_centers,
            'y_centers': y_centers
        }

    def _add_tracking_eoi_labels(self, ax, eoi_info):
        '''Overlay EOI count labels (red) on the tracking plot.'''
        if not eoi_info or 'values' not in eoi_info:
            return
        values = eoi_info['values']
        if np.sum(values) <= 0:
            return
        if eoi_info.get('is_polar'):
            center_x, center_y = eoi_info['center']
            width = eoi_info['width']
            height = eoi_info['height']
            r_edges = eoi_info['radius_edges']
            theta_edges = eoi_info['angle_edges']
            r_centers = 0.5 * (r_edges[:-1] + r_edges[1:])
            theta_centers = 0.5 * (theta_edges[:-1] + theta_edges[1:])
            for i, r in enumerate(r_centers):
                for j, theta in enumerate(theta_centers):
                    cnt = values[i, j]
                    if cnt <= 0:
                        continue
                    r_clamped = min(r, 1.0)
                    x = center_x + (width / 2.0) * r_clamped * np.cos(theta)
                    y = center_y + (height / 2.0) * r_clamped * np.sin(theta)
                    ax.text(x, y, f"{int(cnt)}", fontsize=7, ha='center', va='center', color='red',
                            bbox=dict(boxstyle='round,pad=0.15', facecolor='white', alpha=0.6, linewidth=0))
        else:
            x_centers = eoi_info['x_centers']
            y_centers = eoi_info['y_centers']
            for i, xc in enumerate(x_centers):
                for j, yc in enumerate(y_centers):
                    cnt = values[i, j]
                    if cnt <= 0:
                        continue
                    ax.text(xc, yc, f"{int(cnt)}", fontsize=7, ha='center', va='center', color='red',
                            bbox=dict(boxstyle='round,pad=0.15', facecolor='white', alpha=0.6, linewidth=0))

    def _calculate_average_pixmap(self, pixmaps):
        '''Average a list of QPixmaps'''
        if not pixmaps: return QPixmap()
        try:
            # Use RGBA8888 for consistent 4-byte alignment to avoid stride/skew issues
            img0 = pixmaps[0].toImage().convertToFormat(QImage.Format_RGBA8888)
            w, h = img0.width(), img0.height()
            ptr = img0.bits()
            ptr.setsize(h * w * 4)
            arr_sum = np.frombuffer(ptr, np.uint8).reshape((h, w, 4)).astype(np.float32)
            
            for i in range(1, len(pixmaps)):
                img = pixmaps[i].toImage().convertToFormat(QImage.Format_RGBA8888)
                ptr = img.bits()
                ptr.setsize(h * w * 4)
                arr_sum += np.frombuffer(ptr, np.uint8).reshape((h, w, 4)).astype(np.float32)
            
            arr_avg = (arr_sum / len(pixmaps)).astype(np.uint8)
            result_img = QImage(arr_avg.data, w, h, w * 4, QImage.Format_RGBA8888)
            return QPixmap.fromImage(result_img.copy())
        except Exception as e:
            print(f"Error averaging pixmaps: {e}")
            return pixmaps[0]
    
    # ------------------------------------------- #  
    
    def progressBar(self, n):
  
        '''
            Reflects progress of frequency map and scaling factor computations
        '''

        n = int(n)
        # Setting geometry to progress bar
        self.bar.setValue(n)
        
    # ------------------------------------------- #  
    
    def updateLabel(self, value): 

        ''' 
            Updates the value of the progress bar
        '''

        self.progressBar_Label.setText(value)
        
    # ------------------------------------------- #  
    
    def runSession(self):
        
        '''
            Error checks user input, and invokes worker thread function 
            to compute maps, graphs and scaling factors.
        '''

        cbutton = self.sender()
        
        # Determine if this is an auto-run/re-render (files already set) or interactive browse
        is_rerender_or_auto = False
        if cbutton is None:
            is_rerender_or_auto = True
        elif hasattr(cbutton, 'text') and cbutton.text() == 'Re-Render':
            is_rerender_or_auto = True
        
        # Prepare error dialog window 
        boolean_check = True
        self.error_dialog = QErrorMessage()
        
        # If speed input only specifies lower bound, set upperbound to default
        if (self.speed_lowerbound != None and self.speed_upperbound == None):
            self.speed_upperbound = 100
            
        # If speed filter text is left blank, set default to 0cms to 100cms
        if self.speed_lowerbound == None and self.speed_upperbound == None: 
            self.speed_lowerbound = 0
            self.speed_upperbound = 100

         # Sheck speed bounds are ascending
        if self.speed_lowerbound != None and self.speed_upperbound != None:
            if self.speed_lowerbound > self.speed_upperbound: 
                self.error_dialog.showMessage('Speed filter range must be ascending. Lower speed first, higher speed after. Ex: 1,5')
                boolean_check = False
        
        # Error checking ppm 
        if self.ppm == None or self.chunk_size == None: 
            self.error_dialog.showMessage('PPM field and/or Chunk Size field is blank, or has a non-numeric input. Please enter appropriate numerics.')
            boolean_check = False
            
        # If all checks pass, and we are not in re-render mode, query user for files.
        if boolean_check: 
            if not is_rerender_or_auto:
                run_flag = self.openFileNamesDialog()
                # If the user did not choose the correct files, do not execute thread.
                if not run_flag:
                    return
                
                # Set and execute initialize_fMap function in worker thread
                self.worker = Worker(initialize_fMap, self.files, self.ppm, self.chunk_size, self.window_type, 
                            self.speed_lowerbound, self.speed_upperbound)
                self.worker.start()
                # Worker thread signaling
                self.worker.signals.image_data.connect(self.setData)
                self.worker.signals.progress.connect(self.progressBar)
                self.worker.signals.text_progress.connect(self.updateLabel)
             
            # If we are in re-render mode
            else: 
                # If no files chosen
                if self.files[0] == self.files[1] == None:
                     if cbutton: # Only show error if triggered by button click
                        self.error_dialog.showMessage("You haven't selected a session yet from Browse file")
                     return

                else: 
                    # Set re-render button back to default once files have been chosen
                    self.render_button.setStyleSheet("background-color : light gray")
                    # Browse button should stay neutral when EEG/EGF is available
                    self.updateBrowseButtonStyle()
                    # Launch worker thread
                    self.worker = Worker(initialize_fMap, self.files, self.ppm, self.chunk_size, self.window_type, 
                                self.speed_lowerbound, self.speed_upperbound)
                    self.worker.start()
                    self.worker.signals.image_data.connect(self.setData)
                    self.worker.signals.progress.connect(self.progressBar)
                    self.worker.signals.text_progress.connect(self.updateLabel)
                
    # ------------------------------------------- #  
    
    def windowChanged(self, value): 

        '''
            Updates window type if user chooses different window. 
            Also invokes color change on re-render button to signal that a re-render is needed.
        '''

        self.window_type = value
        self.render_button.setStyleSheet("background-color : rgb(0, 180,0)")
        
    # ------------------------------------------- #  
    
    def loadEOIFile(self):
        '''Manually load an EOI file via dialog'''
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(self, "Load EOI File", self.active_folder, 
                                                 "CSV/Text Files (*.csv *.txt);;All Files (*)", options=options)
        if file_path:
            self.eoi_file_arg = file_path
            self.processEOIs()
            # Refresh current plot
            if self.chunk_index is not None:
                self.sliderChanged(self.chunk_index)
            QMessageBox.information(self, "EOIs Loaded", f"Loaded EOIs from {os.path.basename(file_path)}")

    # ------------------------------------------- #  
    
    def processEOIs(self):
        '''Process EOI file and map to tracking segments'''
        self.eoi_segments = {}
        if not self.eoi_file_arg or not os.path.exists(self.eoi_file_arg):
            return

        try:
            import re
            def extract_numbers(tokens):
                nums = []
                for tok in tokens:
                    # Split by common delimiters inside token as well
                    for sub in re.split(r"[\s,;]+", tok.strip()):
                        if not sub:
                            continue
                        try:
                            nums.append(float(sub))
                        except ValueError:
                            # Drop non-numeric (headers like 'Start(ms)')
                            pass
                return nums

            eois = []
            with open(self.eoi_file_arg, 'r', encoding='utf-8', errors='ignore') as f:
                for raw_line in f:
                    line = raw_line.strip()
                    if not line:
                        continue
                    # Skip header/comment lines
                    if line.startswith('#') or re.match(r"^[A-Za-z]", line):
                        continue
                    # Tokenize by broad delimiters
                    tokens = re.split(r"[\t,;]+|\s{2,}", line)
                    nums = extract_numbers(tokens)
                    if len(nums) < 2:
                        continue
                    # Choose start/stop: if 3+, assume first is ID then next two are times
                    if len(nums) >= 3:
                        start_val, stop_val = nums[1], nums[2]
                    else:
                        start_val, stop_val = nums[0], nums[1]
                    # Detect units: if values look like ms compared to session duration, convert
                    session_sec = None
                    try:
                        if self.pos_t is not None and len(self.pos_t):
                            session_sec = float(self.pos_t[-1])
                        elif hasattr(self, 'binned_data') and self.binned_data and 'duration' in self.binned_data:
                            session_sec = float(self.binned_data['duration'])
                    except Exception:
                        session_sec = None
                    if session_sec and (start_val > session_sec * 5 or stop_val > session_sec * 5):
                        # Heuristic: values far larger than session seconds → treat as ms
                        start_val /= 1000.0
                        stop_val /= 1000.0
                    eois.append((start_val, stop_val))
            
            if not eois or self.tracking_data is None:
                return
            
            # Handle 3-tuple (with time chunks) or 2-tuple (legacy)
            if len(self.tracking_data) == 3:
                x_chunks, y_chunks, t_chunks = self.tracking_data
                
                for i, (chunk_x, chunk_y, chunk_t) in enumerate(zip(x_chunks, y_chunks, t_chunks)):
                    if len(chunk_t) == 0: continue
                    
                    for start, stop in eois:
                        # Find indices where time is within EOI
                        idx_start = np.searchsorted(chunk_t, start)
                        idx_end = np.searchsorted(chunk_t, stop)
                        
                        if idx_end > idx_start:
                            if i not in self.eoi_segments: self.eoi_segments[i] = []
                            self.eoi_segments[i].append((chunk_x[idx_start:idx_end], chunk_y[idx_start:idx_end]))
            else:
                print("Warning: Legacy tracking data format (no time chunks). Cannot map EOIs accurately.")
                    
            print(f"Processed {len(eois)} EOIs into segments.")
            
        except Exception as e:
            print(f"Error processing EOIs: {e}")

    # ------------------------------------------- #  
    
    def setData(self, data):

        '''
            Acquire and set references from returned worker thread data.
            For details on the nature of return value, check initialize_fMap.py 
            return description. 
        '''

        # Support both legacy (6-tuple) and new (7-tuple with binned_data)
        self.freq_dict = data[0]
        self.plot_data = data[1]
        self.images = self.freq_dict[self.frequencyBand]
        self.pos_t = data[2]
        self.scaling_factor_crossband = data[3]
        self.chunk_powers_data = data[4]
        self.tracking_data = data[5]
        self.binned_data = data[6] if len(data) > 6 else None
        
        # Calculate global bounds for tracking plot
        if self.tracking_data:
            try:
                all_x = np.concatenate(self.tracking_data[0])
                all_y = np.concatenate(self.tracking_data[1])
                if len(all_x) > 0:
                    min_x, max_x = np.min(all_x), np.max(all_x)
                    min_y, max_y = np.min(all_y), np.max(all_y)
                    self.data_bounds = (min_x, max_x, min_y, max_y)
                    
                    w = max_x - min_x
                    h = max_y - min_y
                    pad_x = w * 0.1 if w > 0 else 1.0
                    pad_y = h * 0.1 if h > 0 else 1.0
                    self.track_xlim = (min_x - pad_x, max_x + pad_x)
                    self.track_ylim = (min_y - pad_y, max_y + pad_y)
            except Exception as e:
                print(f"Error calculating tracking bounds: {e}")
        
        # Process EOIs now that we have tracking data
        self.processEOIs()
        
        # Update tracking label with arena shape if available
        if len(data) > 7:
            self.tracking_Label.setText(f"Animal tracking<br><span style='font-size:12pt'>{data[7]}</span>")
            self.arena_shape = data[7]
        else:
            self.tracking_Label.setText("Animal tracking")
            self.arena_shape = "Unknown"
        
        # Update the binned analysis window if it's open
        if self.binned_analysis_window is not None and self.binned_analysis_window.isVisible():
            self.binned_analysis_window.updateBinnedData(self.binned_data, self.files, self.active_folder, self.eoi_segments, self.tracking_data)
        
        # Slider limits
        self.slider.setMinimum(0)
        self.slider.setMaximum( len(self.images)-1 )
        self.slider.setSingleStep(1)
        
        # Initialize power display with first chunk
        self.chunk_index = 0
        self.updatePowerDisplay(0)
        
        print("Data loaded")

    # ------------------------------------------- #
    
    def openBinnedAnalysisWindow(self):
        '''Open the separate binned analysis studio window'''
        if not hasattr(self, 'binned_data') or self.binned_data is None:
            QMessageBox.information(
                self, 
                'Binned Analysis Studio', 
                'No binned data available yet. Please run a session first.'
            )
            return
        
        # Create or show existing window
        if self.binned_analysis_window is None or not self.binned_analysis_window.isVisible():
            self.binned_analysis_window = BinnedAnalysisWindow(
                parent=self,
                binned_data=self.binned_data,
                files=self.files,
                active_folder=self.active_folder,
                eoi_segments=self.eoi_segments,
                tracking_data=self.tracking_data
            )
        
        self.binned_analysis_window.show()
        self.binned_analysis_window.raise_()
        self.binned_analysis_window.activateWindow()

    # ------------------------------------------- #
    
    def showBinnedAnalysis(self):
        '''Generate and display 4x4 binned analysis heatmap in a modal panel.'''
        if not hasattr(self, 'binned_data') or self.binned_data is None:
            QMessageBox.information(self, 'Binned Analysis', 'No binned analysis available yet. Please run a session first.')
            return
        try:
            # Save visualization to a temporary path in the active folder
            out_dir = self.active_folder or os.getcwd()
            base_name = os.path.splitext(os.path.basename(self.files[1] or 'output'))[0]
            viz_path = os.path.join(out_dir, f"{base_name}_binned_heatmap.png")
            visualize_binned_analysis(self.binned_data, save_path=viz_path)
            # Show in a modal dialog with responsive sizing
            dlg = QDialog(self)
            dlg.setWindowTitle('4x4 Binned Analysis Heatmap')
            vbox = QVBoxLayout(dlg)
            
            # Create a scroll area for responsive sizing
            scroll = QScrollArea(dlg)
            scroll.setWidgetResizable(True)
            img_label = QLabel()
            pix = QPixmap(viz_path)
            img_label.setPixmap(pix)
            img_label.setScaledContents(False)
            scroll.setWidget(img_label)
            vbox.addWidget(scroll)
            
            # Get available screen geometry and size dialog to fit within 70% of screen
            screen_geometry = QDesktopWidget().availableGeometry()
            max_width = int(screen_geometry.width() * 0.7)
            max_height = int(screen_geometry.height() * 0.7)
            # Ensure minimum reasonable size
            dlg_width = min(max_width, max(600, pix.width() + 50))
            dlg_height = min(max_height, max(400, pix.height() + 50))
            dlg.resize(dlg_width, dlg_height)
            dlg.exec_()
        except Exception as e:
            QMessageBox.warning(self, 'Binned Analysis', f'Failed to render binned analysis: {str(e)}')

    # ------------------------------------------- #
    
    def exportBinnedCsvs(self):
        '''Re-export binned CSVs on demand via the Options menu.'''
        if not hasattr(self, 'binned_data') or self.binned_data is None:
            QMessageBox.information(self, 'Export Binned CSVs', 'No binned analysis available yet. Please run a session first.')
            return
        try:
            out_dir = self.active_folder or os.getcwd()
            base_name = os.path.splitext(os.path.basename(self.files[1] or 'output'))[0]
            output_prefix = os.path.join(out_dir, f"{base_name}_binned")
            pos_x_chunks = self.tracking_data[0] if self.tracking_data else None
            pos_y_chunks = self.tracking_data[1] if self.tracking_data else None
            export_binned_analysis_to_csv(self.binned_data, output_prefix, pos_x_chunks=pos_x_chunks, pos_y_chunks=pos_y_chunks, eoi_segments=self.eoi_segments)
            QMessageBox.information(self, 'Export Binned CSVs', f'Exported binned data to:\n{output_prefix}_*.xlsx / *.csv')
        except Exception as e:
            QMessageBox.warning(self, 'Export Binned CSVs', f'Failed to export binned CSVs: {str(e)}')
        
# =========================================================================== #         

def main(): 
    
    '''
        Main function invokes application start.
        Accepts optional command-line arguments for EEG file path and PPM value.
    '''
    
    app = QApplication(sys.argv)
    
    # Check if an EEG file was passed as command-line argument
    eeg_file = None
    ppm = None
    eoi_file = None
    if len(sys.argv) > 1:
        eeg_file = sys.argv[1]
    if len(sys.argv) > 2:
        ppm = sys.argv[2]
    if len(sys.argv) > 3:
        eoi_file = sys.argv[3]
    
    screen = frequencyPlotWindow(eeg_file, ppm, eoi_file)
    screen.show()
    sys.exit(app.exec_())
    
if __name__=="__main__":
    main()
